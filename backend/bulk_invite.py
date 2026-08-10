# -*- coding: utf-8 -*-
"""Массовое приглашение сотрудников по районным спискам (xlsx).

Каждый xlsx-файл — один район, формат как в присланных списках:
  A1: "<Район> — N чел."
  ...
  строка-заголовок: ФИО | Логин | Роль | Телефон
  далее — строки сотрудников; Роль: "Инспектор" | "Проверяющий" (| "Админ")

⚠ Списки содержат персональные данные (ФИО, телефон) — держать только на
сервере (--xlsx-dir), НЕ коммитить в git. После рассылки ссылок исходники
можно удалить.

Работает через тот же HTTP API, что и ручное создание приглашения в
интерфейсе (POST /api/v1/auth/invites) под ролью admin — бизнес-логика и
проверка уникальности логина не дублируются. Идемпотентно: если логин уже
занят, сервер вернёт 409 — строка помечается «уже есть» и пропускается,
повторный запуск ничего не задваивает и не пересоздаёт.

По умолчанию — dry-run: только парсинг, локальные проверки данных
(дубли логинов, неизвестные роли/районы) и вход под админом для сверки
списка районов; сами приглашения не создаются. Рассылка — с флагом --apply.

Запуск на сервере (тот же docker-network, что и у работающего api):
  dry-run: docker compose -f docker-compose.prod.yml run --rm \
             -v /opt/jirajura/rosters:/rosters:ro api python bulk_invite.py \
             --admin-login admin --admin-password '...' \
             --xlsx-dir /rosters --out /rosters/result.csv
  апплай:  тот же вызов + --apply
"""
import argparse
import csv
import glob
import json
import os
import re
import sys
import urllib.error
import urllib.request

ROLE_MAP = {"Инспектор": "inspector", "Проверяющий": "reviewer", "Админ": "admin"}


def api_call(base_url, method, path, token=None, body=None):
    url = base_url.rstrip("/") + path
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Content-Type", "application/json")
    if token:
        req.add_header("Authorization", f"Bearer {token}")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return resp.status, json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        try:
            return e.code, json.loads(e.read().decode())
        except Exception:
            return e.code, {"detail": str(e)}


def parse_roster(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Sheet" not in wb.sheetnames:
        return None
    rows = list(wb["Sheet"].iter_rows(values_only=True))
    if not rows or not rows[0] or not rows[0][0]:
        return None
    title = str(rows[0][0])
    m = re.match(r"^(.+?)\s*—\s*(\d+)\s*чел", title)
    if not m:
        return None
    district = m.group(1).strip()
    declared = int(m.group(2))
    start = None
    for i, r in enumerate(rows):
        if r and r[0] == "ФИО":
            start = i + 1
            break
    if start is None:
        return None
    people = []
    for r in rows[start:]:
        if not r or not r[0]:
            continue
        fio, login, role, phone = (list(r) + [None, None, None, None])[:4]
        people.append({
            "fio": str(fio).strip() if fio else None,
            "login": str(login).strip() if login else None,
            "role_ru": str(role).strip() if role else None,
            "phone": str(phone).strip() if phone else None,
        })
    return {"district": district, "declared": declared, "people": people, "file": path}


def main():
    p = argparse.ArgumentParser(description="Массовое приглашение сотрудников по районным спискам")
    p.add_argument("--base-url", default="http://api:8000/api/v1")
    p.add_argument("--admin-login", required=True)
    p.add_argument("--admin-password", required=True)
    p.add_argument("--xlsx-dir", required=True)
    p.add_argument("--out", required=True,
                    help="CSV с результатом: район;ФИО;логин;роль;телефон;статус;ссылка")
    p.add_argument("--site-url", default="https://journal.yuvibot2.xyz:8443",
                    help="только для сборки ссылки в CSV")
    p.add_argument("--apply", action="store_true")
    args = p.parse_args()

    files = sorted(glob.glob(os.path.join(args.xlsx_dir, "*.xlsx")))
    rosters = [r for r in (parse_roster(f) for f in files) if r]
    if not rosters:
        print(f"ОШИБКА: в {args.xlsx_dir} не найдено файлов-росписей "
              f"(*.xlsx с заголовком «<Район> — N чел.»)", file=sys.stderr)
        sys.exit(1)

    total = sum(len(r["people"]) for r in rosters)
    print(f"Файлов-росписей: {len(rosters)}, сотрудников всего: {total}")

    # локальные проверки данных
    seen_logins = {}
    row_status = {}  # id(person) -> причина пропуска, если есть
    for r in rosters:
        if len(r["people"]) != r["declared"]:
            print(f"⚠ [{r['district']}] в заголовке заявлено {r['declared']}, "
                  f"фактически строк {len(r['people'])}")
        for person in r["people"]:
            if not person["login"]:
                row_status[id(person)] = "пустой логин"
                continue
            if person["role_ru"] not in ROLE_MAP:
                row_status[id(person)] = f"неизвестная роль «{person['role_ru']}»"
                continue
            key = person["login"].lower()
            if key in seen_logins:
                row_status[id(person)] = f"дубль логина (первый — {seen_logins[key]})"
            else:
                seen_logins[key] = f"[{r['district']}] {person['fio']}"

    dup_or_bad = sum(1 for v in row_status.values())
    if dup_or_bad:
        print(f"\n⚠ Проблемных строк (будут пропущены): {dup_or_bad}")
        shown = 0
        for r in rosters:
            for person in r["people"]:
                reason = row_status.get(id(person))
                if reason and shown < 30:
                    print(f"   [{r['district']}] {person['fio']} (логин {person['login']}): {reason}")
                    shown += 1

    # вход под админом + список районов (нужен даже для dry-run, чтобы
    # заранее найти районы из файлов, которых нет в БД)
    status, body = api_call(args.base_url, "POST", "/auth/login", body={
        "login": args.admin_login, "password": args.admin_password})
    if status != 200:
        print(f"ОШИБКА входа под админом: HTTP {status} {body}", file=sys.stderr)
        sys.exit(1)
    token = body["access_token"]
    if body["user"]["role"] != "admin":
        print(f"ОШИБКА: пользователь {args.admin_login} не админ "
              f"(роль {body['user']['role']}) — приглашения создавать не может", file=sys.stderr)
        sys.exit(1)
    print(f"Вход под админом выполнен.")

    status, districts = api_call(args.base_url, "GET", "/districts/", token=token)
    if status != 200:
        print(f"ОШИБКА получения районов: HTTP {status} {districts}", file=sys.stderr)
        sys.exit(1)
    district_by_name = {d["name"].strip(): d["id"] for d in districts}

    # Уже отправленные, ещё не завершённые приглашения — чтобы при повторном
    # запуске (например, с исправленным логином в исходнике) не создавать
    # ВТОРОЕ приглашение тому же человеку под другим логином вместо того,
    # чтобы понять "ему уже отправляли, проверьте вручную". Именно так
    # раньше набегали дубли вида ChotchaevRB/ChotchaevRB1 на один и тот же
    # человек — сервер видел разные логины и не мог сам их сопоставить.
    status, pending = api_call(args.base_url, "GET", "/auth/invites/pending", token=token)
    if status != 200:
        print(f"ОШИБКА получения списка приглашений: HTTP {status} {pending}", file=sys.stderr)
        sys.exit(1)
    pending_by_name = {
        (inv["district_id"], inv["full_name"].strip().lower())
        for inv in pending
    }

    unknown_districts = sorted({r["district"] for r in rosters
                                 if r["district"] not in district_by_name})
    if unknown_districts:
        print(f"\n⚠ Районы из файлов, не найденные в БД (эти файлы целиком пропущены): "
              f"{unknown_districts}")

    rows_out = []
    created = already = failed = skipped = 0
    for r in rosters:
        dist_id = district_by_name.get(r["district"])
        for person in r["people"]:
            reason = row_status.get(id(person))
            if reason is None and dist_id is None:
                reason = "район не найден в БД"
            if reason is None and person["fio"] and (dist_id, person["fio"].strip().lower()) in pending_by_name:
                reason = "уже есть неиспользованное приглашение под другим логином — проверьте вручную"
            if reason:
                rows_out.append([r["district"], person["fio"], person["login"],
                                  person["role_ru"], person["phone"],
                                  f"ПРОПУЩЕНО: {reason}", ""])
                skipped += 1
                continue

            role = ROLE_MAP[person["role_ru"]]
            if not args.apply:
                rows_out.append([r["district"], person["fio"], person["login"],
                                  person["role_ru"], person["phone"], "К СОЗДАНИЮ (dry-run)", ""])
                continue

            status, resp = api_call(args.base_url, "POST", "/auth/invites", token=token, body={
                "login": person["login"], "full_name": person["fio"],
                "role": role, "district_id": dist_id,
            })
            if status == 200:
                link = f"{args.site_url.rstrip('/')}/register/{resp['token']}"
                rows_out.append([r["district"], person["fio"], person["login"],
                                  person["role_ru"], person["phone"], "СОЗДАНО", link])
                created += 1
            elif status == 409:
                rows_out.append([r["district"], person["fio"], person["login"],
                                  person["role_ru"], person["phone"],
                                  "УЖЕ ЗАРЕГИСТРИРОВАН/ПРИГЛАШЁН", ""])
                already += 1
            else:
                rows_out.append([r["district"], person["fio"], person["login"],
                                  person["role_ru"], person["phone"],
                                  f"ОШИБКА HTTP {status}: {resp}", ""])
                failed += 1

    with open(args.out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Район", "ФИО", "Логин", "Роль", "Телефон", "Статус", "Ссылка"])
        w.writerows(rows_out)

    print(f"\nПропущено (проблемы в данных/районе): {skipped}")
    if not args.apply:
        print(f"К созданию годится: {total - skipped}")
        print("\nЭто был dry-run — приглашения не создавались. Для рассылки добавьте --apply.")
    else:
        print(f"Создано новых приглашений:            {created}")
        print(f"Уже были (логин занят):               {already}")
        print(f"Ошибок:                                {failed}")
    print(f"Отчёт сохранён: {args.out}")


if __name__ == "__main__":
    main()
