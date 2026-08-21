"""Регрессия: /reports/export.xlsx содержит лист "Обзор" (эталонный формат
из generate_summary_report.py). Лист "Регистрация" и все связанные с
регистрацией пользователей цифры намеренно убраны из отчёта — выгрузка
теперь про то, что происходит с площадками и замечаниями (в первую
очередь — устранение), а не про учёт заведённых аккаунтов. См.
app/routers/reports.py: export_xlsx, блок "Топ районов по устранению
замечаний" (считает по IssueStatusHistory.new_status == 'closed')."""
import io
import os
import uuid

import psycopg2
import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

SYNC_DB_URL = os.environ["DATABASE_URL"].replace("postgresql+asyncpg://", "postgresql://")


def _exec(sql, params=None):
    conn = psycopg2.connect(SYNC_DB_URL)
    conn.autocommit = True
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or {})
    finally:
        conn.close()


@pytest.mark.asyncio
async def test_export_overview_sheet_has_no_registration_content(client: AsyncClient, admin_headers):
    r = await client.get("/api/v1/reports/export.xlsx", headers=admin_headers)
    assert r.status_code == 200, r.text

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames[0] == "Обзор"
    assert "Регистрация" not in wb.sheetnames

    ov = wb["Обзор"]
    ov_text = "\n".join(str(c.value) for row in ov.iter_rows() for c in row if c.value is not None)
    assert "Ключевые цифры" in ov_text
    assert "Состав отчёта" in ov_text
    assert "Регистрация" not in ov_text
    assert "Закрыто замечаний за период" in ov_text


@pytest.mark.asyncio
async def test_export_overview_has_top_districts_by_closures(client: AsyncClient, admin_headers):
    r = await client.get("/api/v1/districts/", headers=admin_headers)
    district_id = r.json()[0]["id"]

    me = await client.get("/api/v1/auth/me", headers=admin_headers)
    admin_id = me.json()["id"]

    invite = await client.post("/api/v1/auth/invites", json={
        "login": "ClosuresTestInspector", "full_name": "Закрывающий Инспектор",
        "role": "inspector", "district_id": district_id,
    }, headers=admin_headers)
    complete = await client.post(
        f"/api/v1/auth/invites/{invite.json()['token']}/complete", json={"password": "Closures12345"},
    )
    inspector_headers = {"Authorization": f"Bearer {complete.json()['access_token']}"}

    court_id, site_id = str(uuid.uuid4()), str(uuid.uuid4())
    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES (%(c)s, %(d)s, %(n)s)",
        {"c": court_id, "d": district_id, "n": "Двор для закрытий"},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "(%(s)s, %(c)s, 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((44.0 55.8,44.001 55.8,44.001 55.801,44.0 55.801,44.0 55.8))', 4326), true)",
        {"s": site_id, "c": court_id},
    )
    start = await client.post("/api/v1/inspections/", json={"site_id": site_id, "type": "regular"}, headers=inspector_headers)
    insp_id = start.json()["id"]

    issue_id = str(uuid.uuid4())
    _exec(
        "INSERT INTO issues (id, inspection_id, site_id, category_id, title, status, created_by) VALUES "
        "(%(i)s, %(insp)s, %(s)s, (SELECT id FROM issue_categories WHERE name='Прочее'), "
        "'Тестовое замечание для закрытий', 'closed', %(u)s)",
        {"i": issue_id, "insp": insp_id, "s": site_id, "u": admin_id},
    )
    _exec(
        "INSERT INTO issue_status_history (id, issue_id, old_status, new_status, changed_by) VALUES "
        "(%(h)s, %(i)s, 'fixed', 'closed', %(u)s)",
        {"h": str(uuid.uuid4()), "i": issue_id, "u": admin_id},
    )

    r = await client.get("/api/v1/reports/export.xlsx", headers=admin_headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ov = wb["Обзор"]
    ov_text = "\n".join(str(c.value) for row in ov.iter_rows() for c in row if c.value is not None)
    assert "Районы: закрыто замечаний" in ov_text

    bar_charts = [c for c in ov._charts if type(c).__name__ == "BarChart"]
    titles = [c.title.tx.rich.p[0].r[0].t for c in bar_charts]
    assert any("Топ районов по устранению замечаний" in t for t in titles)
    closures_chart = next(c for c, t in zip(bar_charts, titles) if "Топ районов по устранению замечаний" in t)
    assert closures_chart.x_axis.axPos == "b"  # см. test_reports_export_charts.py — иначе Excel не рисует подписи районов
