"""Обращения с публичной веб-формы (/feedback) — без авторизации на приём,
только очередь для ручного разбора admin/reviewer. См. FeedbackReport в
models.py — сознательно отдельная сущность от Issue: это жалоба
гражданина/сотрудника, а не находка инспектора при обходе.
"""
import io
import os
import uuid as _uuid
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response, UploadFile, File, Request
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings
from app.database import get_db
from app.models import FeedbackReport, FeedbackAttachment, User
from app.schemas import (
    FeedbackReportCreate, FeedbackReportUpdate, FeedbackReportOut,
    FeedbackReportListOut, FeedbackAttachmentOut,
)
from app.services.permissions import require_role
from app.services.rate_limit import RateLimiter
from app.services.audit import log_action

router = APIRouter()

# Публичный эндпоинт загрузки вложений: лимит числа файлов на обращение снят,
# поэтому защищаем диск от флуда на уровне IP (число загрузок + суммарный
# объём за окно). Синглтон на уровне процесса — на проде api один воркер.
_upload_limiter = RateLimiter(
    max_requests=settings.FEEDBACK_ATTACHMENT_MAX_PER_WINDOW,
    window_seconds=settings.FEEDBACK_ATTACHMENT_RATE_WINDOW_SECONDS,
    max_bytes=settings.FEEDBACK_ATTACHMENT_MAX_BYTES_PER_WINDOW,
)

# Создание обращения — тоже публичный эндпоинт: без лимита спамер плодит
# строки в БД бесконечно (вложения от этого уже защищены _upload_limiter).
_submit_limiter = RateLimiter(
    max_requests=settings.FEEDBACK_SUBMIT_MAX_PER_WINDOW,
    window_seconds=settings.FEEDBACK_SUBMIT_RATE_WINDOW_SECONDS,
)


def _client_ip(request: Request) -> str:
    """X-Real-IP выставляет nginx как $remote_addr (не берёт заголовок от
    клиента), поэтому он точнее X-Forwarded-For, который uvicorn кладёт в
    request.client.host и который клиент может подделать первым значением.
    """
    x_real_ip = request.headers.get("x-real-ip")
    if x_real_ip:
        return x_real_ip.strip()
    if request.client and request.client.host:
        return request.client.host
    return "unknown"


def _require_upload_rate_limit(request: Request) -> str:
    key = _client_ip(request)
    allowed, retry_after = _upload_limiter.check(key)
    if not allowed:
        raise HTTPException(
            429,
            "Слишком много вложений подряд — попробуйте позже",
            headers={"Retry-After": str(retry_after)},
        )
    return key


def _require_submit_rate_limit(request: Request) -> None:
    allowed, retry_after = _submit_limiter.check(_client_ip(request))
    if not allowed:
        raise HTTPException(
            429,
            "Слишком много обращений подряд — попробуйте позже",
            headers={"Retry-After": str(retry_after)},
        )

TYPE_LABELS_RU = {"site": "Площадка", "app": "Приложение", "other": "Другое"}
STATUS_LABELS_RU = {"new": "Новое", "in_review": "В работе", "resolved": "Решено", "dismissed": "Отклонено"}

_STATUSES = ("new", "in_review", "resolved", "dismissed")

# Публичный эндпоинт без авторизации — белый список расширений вместо
# чёрного, чтобы нельзя было залить исполняемый файл на сервер.
_ALLOWED_EXTENSIONS = {
    "jpg", "jpeg", "png", "heic", "webp",
    "pdf", "doc", "docx", "xls", "xlsx", "csv", "txt",
}
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "uploads")


def _attachment_to_out(a: FeedbackAttachment) -> FeedbackAttachmentOut:
    return FeedbackAttachmentOut(
        id=a.id, url=f"/uploads/{a.storage_path}",
        original_filename=a.original_filename, content_type=a.content_type,
        size_bytes=a.size_bytes, created_at=a.created_at,
    )


def _report_to_out(r: FeedbackReport) -> FeedbackReportOut:
    return FeedbackReportOut(
        id=r.id, report_type=r.report_type, full_name=r.full_name, phone=r.phone,
        location_text=r.location_text, message=r.message, status=r.status,
        admin_comment=r.admin_comment, created_at=r.created_at, resolved_at=r.resolved_at,
        attachments=[_attachment_to_out(a) for a in (r.attachments or [])],
    )


@router.post("/", response_model=FeedbackReportOut, status_code=201)
async def submit_feedback(
    data: FeedbackReportCreate,
    db: AsyncSession = Depends(get_db),
    _rate_limit: None = Depends(_require_submit_rate_limit),
):
    """Публичный эндпоинт — без авторизации, заявитель может быть анонимным."""
    report = FeedbackReport(
        report_type=data.report_type,
        full_name=data.full_name or None,
        phone=data.phone or None,
        location_text=data.location_text or None,
        message=data.message,
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)
    # Не трогаем report.attachments — на свежесозданном объекте это
    # ленивая (expired-после-commit) relationship, и обращение к ней вне
    # eager-load вызвало бы MissingGreenlet в асинхронной сессии; у только
    # что созданного обращения вложений в любом случае ещё нет.
    return FeedbackReportOut(
        id=report.id, report_type=report.report_type, full_name=report.full_name,
        phone=report.phone, location_text=report.location_text, message=report.message,
        status=report.status, admin_comment=report.admin_comment,
        created_at=report.created_at, resolved_at=report.resolved_at, attachments=[],
    )


@router.post("/{report_id}/attachments", response_model=FeedbackAttachmentOut, status_code=201)
async def upload_feedback_attachment(
    report_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    rate_key: str = Depends(_require_upload_rate_limit),
):
    """Публичный — прикрепить фото/файл к уже созданному обращению.
    Без авторизации, как и сама форма: заявитель может быть анонимным.
    """
    report = (await db.execute(
        select(FeedbackReport).where(FeedbackReport.id == report_id)
    )).scalar_one_or_none()
    if not report:
        raise HTTPException(404, "Обращение не найдено")

    ext = file.filename.rsplit(".", 1)[-1].lower() if file.filename and "." in file.filename else ""
    if ext not in _ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Недопустимый тип файла: .{ext or '?'}")

    filename = f"{_uuid.uuid4()}.{ext}"
    rel_path = os.path.join("feedback", filename)
    abs_path = os.path.join(UPLOAD_DIR, rel_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)

    max_bytes = settings.MAX_PHOTO_SIZE_MB * 1024 * 1024
    size = 0
    with open(abs_path, "wb") as f:
        while chunk := await file.read(1024 * 1024):
            size += len(chunk)
            if size > max_bytes:
                f.close()
                os.remove(abs_path)
                raise HTTPException(413, f"Файл больше {settings.MAX_PHOTO_SIZE_MB} МБ")
            if not _upload_limiter.consume_bytes(rate_key, len(chunk)):
                f.close()
                os.remove(abs_path)
                raise HTTPException(
                    429,
                    "Превышен суммарный объём вложений за короткое время — попробуйте позже",
                    headers={"Retry-After": str(settings.FEEDBACK_ATTACHMENT_RATE_WINDOW_SECONDS)},
                )
            f.write(chunk)

    attachment = FeedbackAttachment(
        feedback_report_id=report_id,
        storage_path=rel_path,
        original_filename=file.filename,
        content_type=file.content_type,
        size_bytes=size,
    )
    db.add(attachment)
    await db.commit()
    await db.refresh(attachment)
    return _attachment_to_out(attachment)


@router.get("/", response_model=FeedbackReportListOut)
async def list_feedback(
    status: Optional[str] = Query(None),
    report_type: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    # Только admin: FeedbackReport не привязан к району (нет district_id),
    # поэтому "reviewer" здесь означал бы обращения ВСЕГО округа, а не
    # своего района — шире, чем reviewer видит где-либо ещё в приложении.
    # Фронтенд и так открывает /admin/feedback только admin (App.tsx) —
    # держим бэкенд в соответствии с этим, а не шире.
    current_user: User = Depends(require_role("admin")),
):
    filters = []
    if status:
        filters.append(FeedbackReport.status == status)
    if report_type:
        filters.append(FeedbackReport.report_type == report_type)

    total = (await db.execute(
        select(func.count()).select_from(FeedbackReport).where(*filters)
    )).scalar_one()
    rows = (await db.execute(
        select(FeedbackReport).options(selectinload(FeedbackReport.attachments))
        .where(*filters)
        .order_by(FeedbackReport.status == "new", FeedbackReport.created_at.desc())
    )).scalars().unique().all()
    return FeedbackReportListOut(total=total, items=[_report_to_out(r) for r in rows])


@router.get("/export.xlsx")
async def export_feedback_xlsx(
    status: Optional[str] = Query(None),
    report_type: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("admin")),  # см. list_feedback
):
    """Выгрузка обращений в Excel — для админа/разработчика, покрывает и
    жалобы на площадки, и технические проблемы (report_type='app'), и
    прочее, одним листом с колонкой типа/статуса.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from app.services.xlsx_style import style_header_row, style_data_row, safe_append

    filters = []
    if status:
        filters.append(FeedbackReport.status == status)
    if report_type:
        filters.append(FeedbackReport.report_type == report_type)

    rows = (await db.execute(
        select(FeedbackReport).options(selectinload(FeedbackReport.attachments))
        .where(*filters)
        .order_by(FeedbackReport.created_at.desc())
    )).scalars().unique().all()

    def _fmt_dt(value):
        return value.strftime("%d.%m.%Y %H:%M") if value else ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Обращения"
    headers = [
        "Дата", "Тип", "Статус", "ФИО", "Телефон", "Место / где возникло",
        "Сообщение", "Комментарий администратора", "Вложений", "Дата решения",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for r in rows:
        # safe_append — обращения приходят с публичной формы БЕЗ авторизации
        # и без премодерации (message/full_name/location_text/admin_comment
        # могут начинаться с "=" и т.п.), см. app/services/xlsx_style.py.
        safe_append(ws, [
            _fmt_dt(r.created_at),
            TYPE_LABELS_RU.get(r.report_type, r.report_type),
            STATUS_LABELS_RU.get(r.status, r.status),
            r.full_name or "Аноним",
            r.phone or "",
            r.location_text or "",
            r.message,
            r.admin_comment or "",
            len(r.attachments or []),
            _fmt_dt(r.resolved_at),
        ])
        style_data_row(ws, ws.max_row, len(headers))
    widths = [16, 14, 12, 24, 16, 30, 50, 40, 10, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="feedback_export_{stamp}.xlsx"'},
    )


@router.patch("/{report_id}", response_model=FeedbackReportOut)
async def update_feedback(
    report_id: str,
    data: FeedbackReportUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("admin")),  # см. list_feedback
):
    report = (await db.execute(
        select(FeedbackReport).where(FeedbackReport.id == report_id)
        .options(selectinload(FeedbackReport.attachments))
    )).scalar_one_or_none()
    if not report:
        raise HTTPException(404, "Обращение не найдено")

    if data.status is not None:
        if data.status not in _STATUSES:
            raise HTTPException(400, f"Недопустимый статус: {data.status}")
        report.status = data.status
        if data.status in ("resolved", "dismissed") and report.resolved_at is None:
            report.resolved_at = datetime.now(timezone.utc)
        elif data.status in ("new", "in_review"):
            report.resolved_at = None
    if data.admin_comment is not None:
        report.admin_comment = data.admin_comment

    # Разбор обращения меняет статус/комментарий — фиксируем в журнале аудита,
    # как и остальные админ-действия (user_update/site_assign и т.п.).
    audit_fields = {}
    if data.status is not None:
        audit_fields["status"] = data.status
    if data.admin_comment is not None:
        audit_fields["admin_comment"] = data.admin_comment
    if audit_fields:
        await log_action(db, str(current_user.id), "feedback_update", "feedback_report", report_id, audit_fields)

    await db.commit()
    await db.refresh(report)
    return _report_to_out(report)
