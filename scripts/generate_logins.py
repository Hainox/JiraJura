"""Генерация логинов формата MichurinDD (Фамилия + инициалы, транслит)
по ФИО из Excel-файла и запись новой колонкой.
"""
import openpyxl
import re
import unicodedata

SRC = r"C:\Users\dmitr\Downloads\Учетные записи (1) (1).xlsx"
DST = r"C:\Users\dmitr\Downloads\Учетные записи (1) (1) — с логинами.xlsx"

# ГОСТ-транслитерация (как в текущих логинах: Ходаковская→Khodakovskaya, Юркова→Yurkova)
TRANS = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
    'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
    'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
    'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
    'ы': 'y', 'ь': '', 'ъ': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
}


def translit(name: str) -> str:
    """Кириллица -> латиница по словарю (пробелы и прочее убираем)."""
    out = []
    for ch in name.lower():
        if ch in TRANS:
            out.append(TRANS[ch])
        elif ch.isalpha():
            out.append(ch)
    return ''.join(out)


def login_from_fio(fio: str) -> str:
    """'Мичурин Дмитрий Дмитриевич' -> 'MichurinDD'"""
    parts = [p for p in re.split(r'\s+', fio.strip()) if p]
    if not parts:
        return ''
    surname = translit(parts[0])
    # первая буква имени и отчества — как в БД: Yu, Ya, Sh, Kh (первая
    # заглавная, остальные строчные: KhodakovskayaYuV, SemenovMYu)
    initials = ''
    for part in parts[1:3]:
        first = part[0]
        if first.isalpha():
            initials += translit(first).capitalize()
    return (surname.capitalize() + initials) if initials else surname.capitalize()


wb = openpyxl.load_workbook(SRC)
ws = wb.active

# Находим колонку с ФИО (3-я по структуре) и добавляем «Логин» в 8-ю
FIO_COL = 3
LOGIN_COL = ws.max_column + 1  # 8

ws.cell(row=1, column=LOGIN_COL, value='Логин')

used: dict[str, int] = {}
rows = []
for r in range(2, ws.max_row + 1):
    fio = ws.cell(row=r, column=FIO_COL).value
    if not fio:
        continue
    base = login_from_fio(fio)
    if not base:
        continue
    used[base] = used.get(base, 0) + 1
    login = base if used[base] == 1 else f'{base}{used[base]}'
    rows.append((r, fio, login))
    ws.cell(row=r, column=LOGIN_COL, value=login)

wb.save(DST)

print(f'Сохранеno: {DST}')
print(f'Всего строк с ФИО: {len(rows)}')
print()
print('№ | ФИО | Логин')
for r, fio, login in rows:
    print(f'{r-1:>3} | {fio} | {login}')
