"""Регрессия: лист "Задания" в /reports/export.xlsx — проверяющие просили
отдельный способ быстро увидеть, кто отвечает за площадку и когда там
был последний обход, а не только исторический журнал за период (лист
"Обходы", который эту информацию не даёт в разрезе "прямо сейчас")."""
import io
import os

import psycopg2
import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from PIL import Image

_buf = io.BytesIO()
Image.new("RGB", (16, 16), (90, 140, 200)).save(_buf, format="JPEG")
_TINY_JPEG = _buf.getvalue()

# Прямые INSERT'ы идут через psycopg2 (не через app.database.async_session)
# — раздельные event loop'ы pytest-asyncio на каждый тест конфликтуют с уже
# запущенным asyncpg-пулом того же engine, если его переиспользовать вне
# запроса через ASGI-транспорт (та же причина, что у sync-подключений в
# conftest.py для подготовки БД).
SYNC_DB_URL = os.environ["DATABASE_URL"].replace("postgresql+asyncpg://", "postgresql://")


def _exec(sql, params=None):
    conn = psycopg2.connect(SYNC_DB_URL)
    conn.autocommit = True
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or {})
    finally:
        conn.close()


@pytest.mark.asyncio
async def test_export_includes_assignments_sheet(client: AsyncClient, admin_headers):
    r = await client.get("/api/v1/districts/", headers=admin_headers)
    district_id = r.json()[0]["id"]

    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES "
        "('99999999-9999-9999-9999-999999999991', %(d)s, 'Тестовый двор — Задания')",
        {"d": district_id},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "('99999999-9999-9999-9999-999999999992', "
        "'99999999-9999-9999-9999-999999999991', 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((37.5 55.8,37.501 55.8,37.501 55.801,37.5 55.801,37.5 55.8))', 4326), true)"
    )

    r = await client.get("/api/v1/reports/export.xlsx", headers=admin_headers)
    assert r.status_code == 200, r.text

    wb = load_workbook(io.BytesIO(r.content))
    assert "Задания" in wb.sheetnames
    ws = wb["Задания"]

    headers = [c.value for c in ws[1]]
    assert headers == [
        "Район", "Двор", "Тип площадки", "Назначенный инспектор",
        "Телефон", "Последний обход", "Статус последнего обхода",
    ]

    rows = {row[1]: row for row in ws.iter_rows(min_row=2, values_only=True)}
    assert "Тестовый двор — Задания" in rows
    row = rows["Тестовый двор — Задания"]
    assert row[3] == "Не назначена"
    assert row[5] == "Обхода ещё не было"


@pytest.mark.asyncio
async def test_assignments_sheet_shows_inspector_and_sorts_unvisited_first(client: AsyncClient, admin_headers):
    r = await client.get("/api/v1/districts/", headers=admin_headers)
    district_id = r.json()[0]["id"]

    invite = await client.post("/api/v1/auth/invites", json={
        "login": "AssignTestInspector", "full_name": "Тестова Заданиевна",
        "role": "inspector", "district_id": district_id,
    }, headers=admin_headers)
    token = invite.json()["token"]
    complete = await client.post(f"/api/v1/auth/invites/{token}/complete", json={"password": "Test12345"})
    inspector_id = complete.json()["user"]["id"]

    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES "
        "('88888888-8888-8888-8888-888888888881', %(d)s, 'Двор с обходом')",
        {"d": district_id},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "('88888888-8888-8888-8888-888888888882', "
        "'88888888-8888-8888-8888-888888888881', 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((37.5 55.8,37.501 55.8,37.501 55.801,37.5 55.801,37.5 55.8))', 4326), true)"
    )
    # ещё один двор без единого обхода — должен остаться выше в сортировке
    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES "
        "('88888888-8888-8888-8888-888888888883', %(d)s, 'Двор без обхода')",
        {"d": district_id},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "('88888888-8888-8888-8888-888888888884', "
        "'88888888-8888-8888-8888-888888888883', 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((37.6 55.8,37.601 55.8,37.601 55.801,37.6 55.801,37.6 55.8))', 4326), true)"
    )

    assign = await client.patch(
        "/api/v1/sites/88888888-8888-8888-8888-888888888882/assign",
        json={"inspector_id": inspector_id}, headers=admin_headers,
    )
    assert assign.status_code == 200, assign.text

    inspector_headers = {"Authorization": f"Bearer {complete.json()['access_token']}"}
    start = await client.post("/api/v1/inspections/", json={
        "site_id": "88888888-8888-8888-8888-888888888882", "type": "regular",
    }, headers=inspector_headers)
    assert start.status_code == 200, start.text
    insp_id = start.json()["id"]
    photo = await client.post(
        f"/api/v1/inspections/{insp_id}/photos",
        params={"target_type": "inspection"},
        files={"file": ("general.jpg", _TINY_JPEG, "image/jpeg")},
        headers=inspector_headers,
    )
    assert photo.status_code == 200, photo.text
    finish = await client.patch(f"/api/v1/inspections/{insp_id}", json={"status": "completed"}, headers=inspector_headers)
    assert finish.status_code == 200, finish.text

    r = await client.get("/api/v1/reports/export.xlsx", headers=admin_headers)
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Задания"]
    rows_in_order = list(ws.iter_rows(min_row=2, values_only=True))
    rows_by_courtyard = {row[1]: row for row in rows_in_order}

    visited = rows_by_courtyard["Двор с обходом"]
    assert visited[3] == "Тестова Заданиевна"
    assert visited[6] == "Завершён"
    assert visited[5] != "Обхода ещё не было"

    unvisited = rows_by_courtyard["Двор без обхода"]
    assert unvisited[5] == "Обхода ещё не было"

    # непосещённые площадки требуют внимания в первую очередь — идут раньше
    # посещённых в списке
    idx_visited = next(i for i, row in enumerate(rows_in_order) if row[1] == "Двор с обходом")
    idx_unvisited = next(i for i, row in enumerate(rows_in_order) if row[1] == "Двор без обхода")
    assert idx_unvisited < idx_visited
