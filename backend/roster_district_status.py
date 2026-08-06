# -*- coding: utf-8 -*-
"""Сверка поданного списка сотрудников (xlsx) с фактическим состоянием в БД.

Формат входного файла — как в общем списке по округу: строка-заголовок
"№ | Район | ФИО | Должность | Номер телефона | Ссылки", далее — по
человеку в строке; "Ссылки" — регистрационная ссылка вида
.../register/<token>, выданная через bulk_invite.py (или ручное создание
инвайта).

Для каждой строки достаёт токен из ссылки, сверяет с user_invites по
token_hash и статусу (used_at/expires_at), формирует сводку по районам:
  - подано               — строк в файле для этого района
  - без ссылки            — в файле нет .../register/... (инвайт не создавался)
  - не найдено в БД        — токен из файла не совпадает ни с одним инвайтом
                             (обычно значит: инвайт перевыпущен другим
                             процессом, старая ссылка мертва)
  - зарегистрирован        — used_at IS NOT NULL
  - ожидает регистрации    — инвайт валиден, не использован, не истёк
  - истёк срок             — не использован, expires_at в прошлом

⚠ Входной файл содержит ФИО и телефоны — держать только на сервере,
не коммитить в git. Скрипт read-only, ничего не пишет и не меняет в БД.

Запуск на сервере:
  docker compose -f docker-compose.prod.yml exec api \
    python roster_district_status.py --roster /app/uploads/roster.xlsx \
    --out-pending /app/uploads/pending_by_district.csv
"""
import argparse
import asyncio
import csv
import hashlib
import os
from collections import defaultdict
from datetime import datetime, timezone

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+asyncpg://postgres:postgres@db:5432/sao_inspection")

COLUMNS = ["подано", "без ссылки", "не найдено в БД", "зарегистрирован", "ожидает регистрации", "истёк срок"]


def _hash_token(token: str) -> str:
    return hashlib.sha256(token.encode()).hexdigest()


def parse_roster(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    people = []
    for r in ws.iter_rows(values_only=True, min_row=2):
        if not r or not r[1] or not r[2]:
            continue
        link = str(r[5]).strip() if len(r) > 5 and r[5] else None
        token = link.rstrip("/").rsplit("/register/", 1)[-1] if link and "/register/" in link else None
        people.append({
            "district": str(r[1]).strip(),
            "fio": str(r[2]).strip(),
            "position": str(r[3]).strip() if len(r) > 3 and r[3] else None,
            "phone": str(r[4]).strip() if len(r) > 4 and r[4] else None,
            "token": token,
        })
    return people


async def classify(db, person):
    d = person["district"]
    if not person["token"]:
        return d, "без ссылки", "без ссылки — инвайт не создавался"

    token_hash = _hash_token(person["token"])
    row = (await db.execute(text(
        "SELECT used_at, expires_at FROM user_invites WHERE token_hash = :h"
    ), {"h": token_hash})).fetchone()

    if row is None:
        return d, "не найдено в БД", "токен не найден в БД (инвайт перевыпущен?)"
    if row.used_at is not None:
        return d, "зарегистрирован", None
    expires_at = row.expires_at
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        return d, "истёк срок", "истёк срок действия ссылки"
    return d, "ожидает регистрации", "ожидает регистрации"


async def main(roster_path, out_pending):
    people = parse_roster(roster_path)
    print(f"В файле: {len(people)} строк\n")

    engine = create_async_engine(DATABASE_URL, echo=False)
    Session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    by_district = defaultdict(lambda: defaultdict(int))
    pending_rows = []

    async with Session() as db:
        for p in people:
            d, status, reason = await classify(db, p)
            by_district[d][status] += 1
            if reason:
                pending_rows.append([d, p["fio"], p["position"], p["phone"], reason])

    await engine.dispose()

    col_w = max(len(c) for c in COLUMNS) + 2
    header = "Район".ljust(22) + "".join(c.rjust(col_w) for c in COLUMNS)
    print(header)
    print("-" * len(header))
    totals = defaultdict(int)
    for d in sorted(by_district):
        row = by_district[d]
        row["подано"] = sum(row.get(c, 0) for c in COLUMNS if c != "подано")
        line = d.ljust(22) + "".join(str(row.get(c, 0)).rjust(col_w) for c in COLUMNS)
        print(line)
        for c in COLUMNS:
            totals[c] += row.get(c, 0)
    print("-" * len(header))
    print("ИТОГО".ljust(22) + "".join(str(totals[c]).rjust(col_w) for c in COLUMNS))

    if out_pending and pending_rows:
        with open(out_pending, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Район", "ФИО", "Должность", "Телефон", "Статус"])
            w.writerows(sorted(pending_rows, key=lambda r: (r[0], r[1])))
        print(f"\nНе зарегистрированные/проблемные строки ({len(pending_rows)}) сохранены в {out_pending}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Сверка поданных списков по районам с фактической регистрацией в БД")
    ap.add_argument("--roster", required=True, help="xlsx с колонками №, Район, ФИО, Должность, Телефон, Ссылки")
    ap.add_argument("--out-pending", default=None, help="куда сохранить CSV незарегистрированных/проблемных строк")
    args = ap.parse_args()
    asyncio.run(main(args.roster, args.out_pending))
