"""1) Обновить 13 логинов в Excel-таблице
2) Разложить по районам: отдельные файлы с ФИО + Логин
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os

SRC = r"C:\Users\dmitr\Downloads\Учетные записи (1) (1) — с логинами.xlsx"
DST = r"C:\Users\dmitr\Downloads\Учетные записи (1) (1) — с логинами.xlsx"
OUT_DIR = r"C:\Users\dmitr\Downloads\По районам"

os.makedirs(OUT_DIR, exist_ok=True)

# ── 13 исправлений: старый_логин → новый_логин ──
FIXES = {
    'BalitskyPA': 'BalitskiyPA',
    'BoymatovBK1': 'BoymatovBK',
    'VasilevskyNS': 'VasilevskiyNS',
    'GnedarevaMN1': 'GnedarevaMN',
    'DzhalavhanovIP': 'DzhalavkhanovIP',
    'KerimovRS2': 'KerimovRSh',
    'KirzhanovAK': 'KirsanovAK',
    'KlevtsovDS1': 'KlevtsovDS',
    'KorsakovaFV': 'KorsakovAV',
    'OsmolovskyAG': 'OsmolovskiyAG',
    'HorchevAM': 'KhorchevAM',
    'ChumichkinaEV1': 'ChumichkinaEV',
    'YurkovaAV1': 'YurkovaAV',
}

wb = openpyxl.load_workbook(SRC)
ws = wb.active

# ── Применяем исправления ──
fixed = 0
LOGIN_COL = 8
for r in range(2, ws.max_row + 1):
    current = (ws.cell(row=r, column=LOGIN_COL).value or '').strip()
    if current in FIXES:
        ws.cell(row=r, column=LOGIN_COL).value = FIXES[current]
        fixed += 1
        print(f'  Исправлено: {current} → {FIXES[current]}')

wb.save(DST)
print(f'\n✅ Исправлено логинов: {fixed}')
print(f'✅ Файл сохранён: {DST}')

# ── Собираем по районам ──
data = defaultdict(list)  # район → [(фио, логин, роль, телефон), ...]
for r in range(2, ws.max_row + 1):
    district = (ws.cell(row=r, column=2).value or '').strip()
    fio = (ws.cell(row=r, column=3).value or '').strip()
    login = (ws.cell(row=r, column=LOGIN_COL).value or '').strip()
    role = (ws.cell(row=r, column=7).value or '').strip()
    phone = str(ws.cell(row=r, column=5).value or '').strip()
    if fio and login:
        data[district].append((fio, login, role, phone))

# ── Стили ──
hdr_font = Font(bold=True, size=12, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2563EB')
hdr_align = Alignment(horizontal='center', vertical='center')
cell_align = Alignment(horizontal='left', vertical='center')
border = Border(left=Side('thin','D1D5DB'), right=Side('thin','D1D5DB'),
                top=Side('thin','D1D5DB'), bottom=Side('thin','D1D5DB'))
alt_fill = PatternFill('solid', fgColor='F8FAFC')

for district, users in sorted(data.items()):
    safe_name = district.replace(' ', '_').replace('/', '_')
    path = os.path.join(OUT_DIR, f'{safe_name}.xlsx')

    wb_d = openpyxl.Workbook()
    ws_d = wb_d.active

    # Заголовок
    ws_d.merge_cells('A1:D1')
    ws_d.cell(row=1, column=1, value=f'{district} — {len(users)} чел.').font = Font(bold=True, size=14, color='111827')

    # Таблица
    headers = ['ФИО', 'Логин', 'Роль', 'Телефон']
    for c, h in enumerate(headers, 1):
        cl = ws_d.cell(row=3, column=c, value=h)
        cl.font = hdr_font; cl.fill = hdr_fill; cl.alignment = hdr_align; cl.border = border

    for i, (fio, login, role, phone) in enumerate(sorted(users, key=lambda x: x[0])):
        r = 4 + i
        vals = [fio, login, role, f'📞 {phone}' if phone and phone != 'None' else '—']
        for c, v in enumerate(vals, 1):
            cl = ws_d.cell(row=r, column=c, value=v)
            cl.alignment = cell_align; cl.border = border
            if i % 2:
                cl.fill = alt_fill

    ws_d.column_dimensions['A'].width = 42
    ws_d.column_dimensions['B'].width = 22
    ws_d.column_dimensions['C'].width = 16
    ws_d.column_dimensions['D'].width = 22

    wb_d.save(path)
    print(f'  ✅ {safe_name}.xlsx — {len(users)} чел.')

print(f'\n🎯 Готово! {len(data)} файлов в: {OUT_DIR}')
