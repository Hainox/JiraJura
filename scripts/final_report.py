"""Финальный отчёт: обновлённая сводка, список незарегистрированных с телефонами, SQL для исправления логинов."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

USERS_FILE = r"C:\Users\dmitr\Downloads\Учетные записи (1) (1) — с логинами.xlsx"
INSP_FILE = r"C:\Users\dmitr\Downloads\журнал_обходов_2026-08-05.xlsx"
OUT_FILE = r"C:\Users\dmitr\Downloads\Сводки_2026-08-05.xlsx"

# ── Стили ──
hdr_font = Font(bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2563EB')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(horizontal='center', vertical='center')
border = Border(left=Side('thin','D1D5DB'), right=Side('thin','D1D5DB'),
                top=Side('thin','D1D5DB'), bottom=Side('thin','D1D5DB'))
green = PatternFill('solid', fgColor='DCFCE7')
red = PatternFill('solid', fgColor='FEE2E2')
yellow = PatternFill('solid', fgColor='FEF3C7')

def hdr(ws, row, n):
    for c in range(1, n+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border

def cell(ws, r, c, v=None, a=cell_align):
    cl = ws.cell(row=r, column=c); cl.alignment = a; cl.border = border
    if v is not None: cl.value = v
    return cl

# ═══ Данные ═══════════════════════════════════════════════════════

# 83 DB-юзера (из админки)
db_users = {}   # login → (fio_login_совпадает?, fio, role, district)
db_by_fio = {}  # fio → login

# 13 расхождений: фио → (db_login, excel_login)
MISMATCHES = {
    'Балицкий Павел Александрович':       ('BalitskyPA',      'BalitskiyPA'),
    'Бойматов Бехзод Курбанович':         ('BoymatovBK1',     'BoymatovBK'),
    'Василевский Никита Сергеевич':       ('VasilevskyNS',    'VasilevskiyNS'),
    'Гнедарева Марина Николаевна':        ('GnedarevaMN1',    'GnedarevaMN'),
    'Джалавханов Ибрагим Пируллахович':   ('DzhalavhanovIP',  'DzhalavkhanovIP'),
    'Керимов Радж Шахин Оглы':            ('KerimovRS2',      'KerimovRSh'),
    'Кирсанов Андрей Константинович':     ('KirzhanovAK',     'KirsanovAK'),
    'Клевцов Денис Сергеевич':            ('KlevtsovDS1',     'KlevtsovDS'),
    'Корсаков Алексей Васильеви':         ('KorsakovaFV',     'KorsakovAV'),
    'Осмоловский Алексей Геннадьевич':    ('OsmolovskyAG',    'OsmolovskiyAG'),
    'Хорчев Анатолий Михайлович':         ('HorchevAM',       'KhorchevAM'),
    'Чумичкина Екатерина Валерьевна':     ('ChumichkinaEV1',  'ChumichkinaEV'),
    'Юркова Арина Владиславовна':         ('YurkovaAV1',      'YurkovaAV'),
}
mismatch_db_logins = set(db for db, _ in MISMATCHES.values())

# DB raw text (тот же что в compare_users.py)
db_raw = """..."""  # placeholder — используем список из MISMATCHES + захардкоженные 63 совпадения

# Проще: прочитаем Excel логины и сверим напрямую
wb_u = openpyxl.load_workbook(USERS_FILE)
ws_u = wb_u.active
invited = {}  # login → (fio, district, role, phone)
for r in range(2, ws_u.max_row + 1):
    fio = (ws_u.cell(row=r, column=3).value or '').strip()
    login = (ws_u.cell(row=r, column=8).value or '').strip()
    district = (ws_u.cell(row=r, column=2).value or '').strip()
    role = (ws_u.cell(row=r, column=7).value or '').strip()
    phone = str(ws_u.cell(row=r, column=5).value or '').strip()
    if login and fio:
        invited[login] = (fio, district, role, phone)

# DB logins из скриншотов (79 активных, без admin/мусора)
db_active_logins_text = """AndreevaNV AsekovAA BalitskyPA BarkovOA BorodinovaMK BoymatovBK1 BratskayaAS
BuzenkovaES BurmistrovVG VarzhinAI VasilevskyNS GevenovLA GnedarevaMN1 GromovVV GurinovaEG
DzhabrailovaEM DzhalavhanovIP EvdokimovaGN ZababurinPI ZaloginaMM ZeynalovSA IsaychevSM KalachevAA
KalashnikovaIV KarpachevAV KerimovRS2 KiperLA KirzhanovAK KlevtsovDS1 KolesyankinaNV KorotkovVI
KorsakovaFV KuznetsovPV KuzminDYu LomakinDA LukashON MagomedovMI MagomedovNN MasyukevichE
MichurinDD MichurinDD1 MichurinDD2 MonakhovVYu MosinSV NaumovAS NefedovKI NovichkovIR
NurmanovR OsmolovskyAG PerepelkinaNV PozhidaevaMV PugachYuN RomanenkoMA SavchenkoVA
SamarinaNV SamokhvalovVYa SafronovEA SemenovMYu SenSA SkuratovIS SokolovSS SpirinaAN
StakhovaSI StepanchukAV SumakovaNA SukhinVS TalanovaAS TarasikVA TokarevaEE
KhodakovskayaYuV HorchevAM ChekalkinaSM CherkashinMV ChilikinaTN ChumichkinaEV1
ShazamovaAR ShkvarinDS ShokirovSh YurkovaAV1""".split()

db_logins = set(db_active_logins_text)
# Добавим mismatch-db логины (на случай если не все в списке)
db_logins |= mismatch_db_logins

# Строим сводку посекундно
district_stats = defaultdict(lambda: {'inv': 0, 'reg': 0, 'mismatch': 0, 'not_reg': []})

for login, (fio, district, role, phone) in invited.items():
    district_stats[district]['inv'] += 1
    if login in db_logins:
        district_stats[district]['reg'] += 1
    elif fio in MISMATCHES:
        district_stats[district]['reg'] += 1
        district_stats[district]['mismatch'] += 1
    else:
        district_stats[district]['not_reg'].append((fio, login, phone))

# Загружаем данные обходов
wb_i = openpyxl.load_workbook(INSP_FILE)
ws_summary = wb_i['Сводка по районам']
dist_insp = {}
for r in range(2, ws_summary.max_row + 1):
    name = ws_summary.cell(row=r, column=1).value or ''
    dist_insp[name] = {
        'sites': ws_summary.cell(row=r, column=2).value or 0,
        'insp': ws_summary.cell(row=r, column=3).value or 0,
        'issues_cr': ws_summary.cell(row=r, column=4).value or 0,
        'issues_open': ws_summary.cell(row=r, column=6).value or 0,
        'overdue': ws_summary.cell(row=r, column=7).value or 0,
    }

# ═══ Запись Excel ════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ── Лист 1: Регистрация ──
ws1 = wb.active
ws1.title = 'Регистрация по районам'
ws1.merge_cells('A1:G1')
ws1.cell(row=1, column=1, value='Регистрация пользователей — САО (реальные данные из БД)').font = Font(bold=True, size=14)

# Итого
total_inv = sum(d['inv'] for d in district_stats.values())
total_reg = sum(d['reg'] for d in district_stats.values())
ws1.merge_cells('A3:G3')
ws1.cell(row=3, column=1, value=f'ВСЕГО: приглашено {total_inv}, зарегистрировано {total_reg} ({round(total_reg/total_inv*100,1) if total_inv else 0}%), не зарег. {total_inv - total_reg}').font = Font(bold=True, size=11, color='2563EB')

headers = ['Район', 'Приглашено', 'Зарег. (точно)', 'Расхожд. логинов', 'Не зарег.', '% охвата', 'Рейтинг']
for c, h in enumerate(headers, 1):
    ws1.cell(row=5, column=c, value=h)
hdr(ws1, 5, 7)

row = 6
for district in sorted(district_stats.keys()):
    d = district_stats[district]
    pct = round(d['reg'] / d['inv'] * 100, 1) if d['inv'] > 0 else 0
    cell(ws1, row, 1, district, Alignment(horizontal='left', vertical='center'))
    cell(ws1, row, 2, d['inv'])
    cell(ws1, row, 3, d['reg'] - d['mismatch'])
    cell(ws1, row, 4, d['mismatch'] if d['mismatch'] else '—')
    cell(ws1, row, 5, d['inv'] - d['reg'])
    cell(ws1, row, 6, f'{pct}%')
    bar = '█' * int(pct/10) + ('░' * (10 - int(pct/10)))
    cell(ws1, row, 7, bar if pct > 0 else '—')
    # Цвет
    fill = green if pct >= 80 else (yellow if pct >= 30 else red)
    for c in range(1, 8):
        ws1.cell(row=row, column=c).fill = fill
    row += 1

for c, w in enumerate([22, 11, 13, 14, 10, 10, 14], 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

# ── Лист 2: Незарегистрированные с телефонами ──
ws2 = wb.create_sheet('Незарегистрированные')
ws2.merge_cells('A1:E1')
ws2.cell(row=1, column=1, value='Незарегистрированные пользователи — для обзвона').font = Font(bold=True, size=14)
ws2.merge_cells('A3:E3')
ws2.cell(row=3, column=1, value=f'{total_inv - total_reg} человек не завершили регистрацию').font = Font(size=11, color='6B7280')

headers2 = ['Район', 'ФИО', 'Логин', 'Телефон', 'Должность']
for c, h in enumerate(headers2, 1):
    ws2.cell(row=5, column=c, value=h)
hdr(ws2, 5, 5)

row = 6
for district in sorted(district_stats.keys()):
    for fio, login, phone in district_stats[district]['not_reg']:
        cell(ws2, row, 1, district, Alignment(horizontal='left', vertical='center'))
        cell(ws2, row, 2, fio, Alignment(horizontal='left', vertical='center'))
        cell(ws2, row, 3, login)
        cell(ws2, row, 4, '📞 ' + phone if phone and phone != 'None' else '—')
        # Должность из исходного Excel
        position = ''
        for l, (f, d, r, p) in invited.items():
            if f == fio:
                position = ''
                break
        # Быстрый поиск по ws_u
        for rx in range(2, ws_u.max_row+1):
            if (ws_u.cell(row=rx, column=8).value or '').strip() == login:
                position = (ws_u.cell(row=rx, column=4).value or '').strip()
                break
        cell(ws2, row, 5, position or '', Alignment(horizontal='left', vertical='center'))
        row += 1

for c, w in enumerate([22, 40, 20, 20, 40], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ── Лист 3: Сводка обходов ──
ws3 = wb.create_sheet('Сводка обходов')
ws3.merge_cells('A1:G1')
ws3.cell(row=1, column=1, value='Сводка обходов — САО (03.08–05.08.2026)').font = Font(bold=True, size=14)

ws_i = wb_i['Обходы']
status_counts = defaultdict(int)
total_ok = 0; total_def = 0
for r in range(2, ws_i.max_row+1):
    st = ws_i.cell(row=r, column=7).value or ''
    status_counts[st] += 1
    total_ok += int(ws_i.cell(row=r, column=10).value or 0)
    total_def += int(ws_i.cell(row=r, column=11).value or 0)
total_insp = sum(status_counts.values())

ws_iss = wb_i['Замечания']
open_issues = 0
for r in range(2, ws_iss.max_row+1):
    if (ws_iss.cell(row=r, column=7).value or '') in ('Открыто', 'Назначено', 'В работе'):
        open_issues += 1

kpi = [
    ('Всего обходов', total_insp, '1F2937'), ('Завершено', status_counts.get('Завершён', 0), '16A34A'),
    ('В процессе', status_counts.get('В процессе', 0), 'CA8A04'), ('С нарушениями', status_counts.get('Есть нарушения', 0), 'EA580C'),
    ('Критических', status_counts.get('Критический', 0), 'DC2626'), ('Пунктов ОК', total_ok, '16A34A'),
    ('Дефектов', wb_i['Нарушения по чек-листу'].max_row - 1, 'DC2626'), ('Замечаний открыто', open_issues, '7C3AED'),
]
for i, (lbl, val, clr) in enumerate(kpi):
    c = i * 2 + 1
    ws3.cell(row=3, column=c, value=lbl).font = Font(size=10, color='6B7280')
    ws3.cell(row=4, column=c, value=val).font = Font(bold=True, size=20, color=clr)
    ws3.column_dimensions[get_column_letter(c)].width = 18

# Таблица по районам (из журнала обходов)
ws3.merge_cells('A7:H7')
ws3.cell(row=7, column=1, value='По районам').font = Font(bold=True, size=12)
headers3 = ['Район', 'Площадок', 'Обходов', 'Замечаний созд.', 'Открыто', 'Просрочено', '% охвата', 'Рег. пользователей']
for c, h in enumerate(headers3, 1):
    ws3.cell(row=8, column=c, value=h)
hdr(ws3, 8, 8)

row = 9
for district in sorted(dist_insp.keys()):
    di = dist_insp[district]
    pct = round(di['insp'] / di['sites'] * 100, 1) if di['sites'] > 0 else 0
    ds = district_stats.get(district, None)
    reg_pct = f"{round(ds['reg']/ds['inv']*100,1)}%" if ds and ds['inv'] > 0 else '—'
    cell(ws3, row, 1, district, Alignment(horizontal='left', vertical='center'))
    cell(ws3, row, 2, di['sites']); cell(ws3, row, 3, di['insp'])
    cell(ws3, row, 4, di['issues_cr']); cell(ws3, row, 5, di['issues_open'])
    cell(ws3, row, 6, di['overdue']); cell(ws3, row, 7, f'{pct}%')
    cell(ws3, row, 8, reg_pct)
    row += 1

wb.save(OUT_FILE)

# ═══ SQL для исправления логинов ════════════════════════════════
print("✅ Файл сохранён:", OUT_FILE)
print(f"\n=== SQL ДЛЯ ИСПРАВЛЕНИЯ 13 ЛОГИНОВ (запустить на сервере) ===\n")
for fio, (db_login, excel_login) in sorted(MISMATCHES.items()):
    print(f"-- {fio}")
    print(f"UPDATE users SET login = '{excel_login}' WHERE login = '{db_login}';")
print()
print(f"Всего приглашено: {total_inv}")
print(f"Зарегистрировано: {total_reg} ({round(total_reg/total_inv*100,1)}%)")
print(f"Не зарегистрировано: {total_inv - total_reg}")
