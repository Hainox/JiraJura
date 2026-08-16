"""Регрессии на публичный submit, валидацию типа, выгрузку и аудит обращений."""
import io

import pytest
from openpyxl import load_workbook

from app.services.rate_limit import RateLimiter


@pytest.mark.asyncio
async def test_feedback_rejects_invalid_report_type(client):
    """Некорректный report_type — 422, а не тихое приведение к 'site'."""
    res = await client.post(
        "/api/v1/feedback/",
        json={"report_type": "bogus", "message": "Проверка валидации типа"},
    )
    assert res.status_code == 422, res.text


@pytest.mark.asyncio
async def test_feedback_submit_rate_limited(client, monkeypatch):
    """Публичное создание обращения тоже лимитируется по IP (защита от спама в БД)."""
    from app.routers import feedback

    monkeypatch.setattr(
        feedback,
        "_submit_limiter",
        RateLimiter(max_requests=3, window_seconds=60),
    )

    for index in range(3):
        res = await client.post(
            "/api/v1/feedback/",
            json={"report_type": "app", "message": f"Сообщение {index}"},
        )
        assert res.status_code == 201, res.text

    rejected = await client.post(
        "/api/v1/feedback/",
        json={"report_type": "app", "message": "Лишнее обращение"},
    )
    assert rejected.status_code == 429, rejected.text
    assert rejected.headers.get("retry-after")


@pytest.mark.asyncio
async def test_feedback_export_xlsx(client, admin_headers):
    """Выгрузка в Excel отдаёт валидный xlsx и уважает фильтр по типу."""
    await client.post(
        "/api/v1/feedback/",
        json={"report_type": "app", "full_name": "Тестовый Гражданин", "message": "Не работает карта"},
    )
    await client.post(
        "/api/v1/feedback/",
        json={"report_type": "site", "full_name": "Другой Гражданин", "message": "Сломаны качели"},
    )

    res = await client.get(
        "/api/v1/feedback/export.xlsx",
        params={"report_type": "app"},
        headers=admin_headers,
    )
    assert res.status_code == 200, res.text
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws["A1"].value == "Дата"
    messages = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
    assert "Не работает карта" in messages
    assert "Сломаны качели" not in messages


@pytest.mark.asyncio
async def test_feedback_update_logs_audit(client, admin_headers):
    """Смена статуса/комментария обращения пишется в журнал аудита."""
    created = await client.post(
        "/api/v1/feedback/",
        json={"report_type": "site", "message": "Сломаны качели во дворе"},
    )
    assert created.status_code == 201, created.text
    report_id = created.json()["id"]

    updated = await client.patch(
        f"/api/v1/feedback/{report_id}",
        json={"status": "in_review", "admin_comment": "Проверим на этой неделе"},
        headers=admin_headers,
    )
    assert updated.status_code == 200, updated.text
    assert updated.json()["status"] == "in_review"
    assert updated.json()["resolved_at"] is None

    from sqlalchemy import select

    from app.database import async_session
    from app.models import AuditLog

    async with async_session() as db:
        rows = (
            await db.execute(
                select(AuditLog).where(
                    AuditLog.entity_type == "feedback_report",
                    AuditLog.entity_id == report_id,
                )
            )
        ).scalars().all()
    assert any(r.action == "feedback_update" for r in rows)
