"""Единый визуальный стандарт для всех Excel-выгрузок приложения (эталон
задан пользователем 11.08.2026 — отформатированная вручную копия
/reports/export.xlsx с жёлтыми заголовками, рамкой и центровкой по всем
ячейкам). Используется во всех /export.xlsx-эндпоинтах, чтобы выгрузки не
расходились по оформлению между собой.
"""
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

HEADER_FILL = PatternFill("solid", fgColor="FFFF00")
_THIN = Side(style="thin")
ALL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_cell(cell, fill: bool = True) -> None:
    """Заголовок таблицы: жирный, жёлтая заливка, рамка, центр, перенос слов.
    fill=False — для вспомогательных мини-таблиц (например, под диаграмму),
    те в эталоне выделены только жирным и рамкой, без жёлтого фона.
    """
    cell.font = Font(bold=True)
    if fill:
        cell.fill = HEADER_FILL
    cell.border = ALL_BORDER
    cell.alignment = CENTER_WRAP


def style_header_row(ws, row_idx: int, ncols: int, fill: bool = True) -> None:
    for col in range(1, ncols + 1):
        style_header_cell(ws.cell(row_idx, col), fill=fill)


def style_data_cell(cell) -> None:
    """Обычная ячейка данных: рамка, центр, перенос слов — без заливки,
    заливку (например, цвет по проценту регистрации) ставит вызывающий код
    поверх после этого вызова.
    """
    cell.border = ALL_BORDER
    cell.alignment = CENTER_WRAP


def style_data_row(ws, row_idx: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        style_data_cell(ws.cell(row_idx, col))


# Excel/openpyxl формула-инъекция (CWE-1236): openpyxl сам детектит "=" в
# начале строкового значения и пишет ячейку как формулу (data_type='f'), а
# не как текст — открытие такого файла в Excel/Google Sheets ИСПОЛНЯЕТ её
# (например =HYPERLINK(...) со ссылкой на фишинг). Во все наши выгрузки
# попадает текст, который вводят пользователи без всякой премодерации
# (замечания инспектора, комментарии чек-листа, анонимные обращения с
# публичной формы) — без этой защиты любой из них мог подложить формулу,
# которая сработает у админа при открытии отчёта.
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _is_risky(value) -> bool:
    return isinstance(value, str) and value[:1] in _FORMULA_TRIGGER_CHARS


def safe_append(ws, row) -> None:
    """ws.append(row), но принудительно помечает как ТЕКСТ (data_type='s')
    любую ячейку строки, чьё значение похоже на формулу — вместо '=...'
    openpyxl закодировал бы её как реальную формулу. Использовать вместо
    ws.append(...) везде, где в строку могут попасть данные, введённые
    пользователем (а не только вычисленные приложением значения)."""
    ws.append(row)
    r = ws.max_row
    for i, v in enumerate(row, start=1):
        if _is_risky(v):
            ws.cell(r, i).data_type = "s"


SPACER_ROW_HEIGHT = 8


def style_merged_label(ws, row_idx: int, ncols: int, header: bool = False, fill: bool = True) -> None:
    """Одноколоночная 'ярлык'-строка (заголовок секции или длинная строка
    текста/сноска), слитая на ncols колонок в один заполненный блок вместо
    узкой рамки в одной колонке рядом с пустым нестилизованным местом.

    Рамку/заливку ставим на КАЖДОЙ покрываемой ячейке, а не только на
    первой: Excel рисует край слитой ячейки по границе именно той
    исходной ячейки, что на этом краю — если стилизована только первая
    (левая) ячейка, у блока не будет рамки с правой стороны.
    """
    for col in range(1, ncols + 1):
        cell = ws.cell(row_idx, col)
        if header:
            style_header_cell(cell, fill=fill)
        else:
            style_data_cell(cell)
    if ncols > 1:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
