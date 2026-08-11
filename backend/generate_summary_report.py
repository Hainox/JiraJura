# -*- coding: utf-8 -*-
"""Сводный отчёт по проекту в формате xlsx (для передачи начальнику) — только
читает БД, ничего не меняет.

Объединяет регистрацию (из production_status_report.py) и выгрузку журнала
(из /reports/export.xlsx) в один файл с добавлением сводной страницы «Обзор»
с ключевыми цифрами и авто-подсказками, на что смотреть в первую очередь.

Листы: Обзор, Регистрация, Сводка по районам, Обходы, Нарушения по
чек-листу, Замечания, Просроченные замечания, Динамика.

Запуск на сервере:
  docker compose -f docker-compose.prod.yml exec api python generate_summary_report.py
  docker compose -f docker-compose.prod.yml cp api:/app/summary_report.xlsx ./summary_report.xlsx

Файл содержит ФИО и телефоны сотрудников — не коммитить в git, скачивать
только на свою машину и не оставлять в публично раздаваемых директориях
(на сервере /app/uploads/ раздаётся наружу через nginx — сюда не пишем).
"""
import asyncio
import os
from collections import defaultdict, Counter
from datetime import datetime, timedelta, timezone

from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+asyncpg://postgres:postgres@db:5432/sao_inspection")

# Те же словари, что в app/routers/reports.py — не импортируем оттуда напрямую,
# чтобы скрипт не тянул за собой весь FastAPI-граф зависимостей (JWT/настройки)
# ради трёх словарей, как и остальные разовые скрипты в этой директории.
INSPECTION_STATUS_RU = {
    "planned": "Запланирован", "in_progress": "В процессе",
    "completed": "Завершён", "issues_found": "Есть нарушения",
    "critical": "Критический",
}
ISSUE_STATUS_RU = {
    "open": "Открыто", "assigned": "Назначено", "in_work": "В работе",
    "fixed": "Устранено", "control": "На контроле", "closed": "Закрыто",
    "overdue": "Просрочено", "revision_needed": "На доработке",
}
CRITICALITY_RU = {
    "low": "Низкая", "medium": "Средняя", "high": "Высокая",
    "critical": "Критическая",
}
OUT_PATH = os.getenv("SUMMARY_REPORT_OUT", "/app/summary_report.xlsx")

# Россия не переходит на летнее время с 2014 — фиксированный UTC+3 стабилен
# круглый год, тот же принцип, что и в backend/app/routers/inspections.py.
MSK = timezone(timedelta(hours=3))


def _fmt_dt(value) -> str:
    return value.strftime("%d.%m.%Y %H:%M") if value else ""


async def fetch_all(db: AsyncSession) -> dict:
    data = {}

    districts = (await db.execute(text(
        "SELECT id, name FROM districts ORDER BY name"
    ))).fetchall()
    data["districts"] = districts

    # ── Регистрация (та же логика, что в production_status_report.py) ──
    pending_rows = (await db.execute(text(
        "SELECT district_id, full_name, login, role "
        "FROM user_invites WHERE used_at IS NULL ORDER BY full_name, created_at DESC"
    ))).fetchall()

    # Один человек может попасть в выборку дважды, если его приглашали
    # повторно под чуть другим логином (опечатка/транслитерация) — дедуп
    # оставляет самую свежую запись (после ORDER BY full_name, created_at
    # DESC первая по имени и есть самая свежая). Важно: дедуп делаем ПОСЛЕ
    # разбивки по районам/ролям, а не по ФИО глобально на весь округ — у
    # full_name нет уникальности в базе, и два разных человека с одинаковым
    # ФИО в разных районах иначе схлопывались бы в одного, а второй молча
    # пропадал бы из отчёта.
    pending_by_district_raw = defaultdict(list)
    pending_admins_raw, pending_okrug_reviewers_raw = [], []
    for r in pending_rows:
        if r.role == "admin":
            pending_admins_raw.append(r)
        elif r.district_id is None:
            pending_okrug_reviewers_raw.append(r)
        else:
            pending_by_district_raw[r.district_id].append(r)

    def _dedup_by_name(rows):
        seen, out = set(), []
        for r in rows:
            if r.full_name in seen:
                continue
            seen.add(r.full_name)
            out.append(r)
        return out

    pending_admins = _dedup_by_name(pending_admins_raw)
    pending_okrug_reviewers = _dedup_by_name(pending_okrug_reviewers_raw)
    pending_by_district = {did: _dedup_by_name(rows) for did, rows in pending_by_district_raw.items()}

    reg_stats = []
    for d in districts:
        registered = (await db.execute(text(
            "SELECT COUNT(*) FROM users WHERE district_id = :did AND is_active"
        ), {"did": d.id})).scalar()
        pending = pending_by_district.get(d.id, [])
        total = registered + len(pending)
        pct = (registered / total) if total else 1.0
        reg_stats.append({
            "name": d.name, "registered": registered, "pending": pending,
            "total": total, "pct": pct,
        })
    reg_stats.sort(key=lambda x: x["pct"])
    data["reg_stats"] = reg_stats
    data["pending_admins"] = pending_admins
    data["pending_okrug_reviewers"] = pending_okrug_reviewers

    total_reg = sum(x["registered"] for x in reg_stats)
    total_all = sum(x["total"] for x in reg_stats)
    data["overall_reg_pct"] = (total_reg / total_all) if total_all else 1.0
    data["total_reg"] = total_reg
    data["total_all_invited"] = total_all

    # ── Работа по районам ──
    work_stats = []
    for d in districts:
        site_ids = [r.id for r in (await db.execute(text(
            "SELECT s.id FROM sites s JOIN courtyards c ON c.id = s.courtyard_id "
            "WHERE c.district_id = :did AND s.is_active"
        ), {"did": d.id})).fetchall()]

        if not site_ids:
            work_stats.append({
                "name": d.name, "sites": 0, "insp_total": 0, "insp_done": 0,
                "iss_total": 0, "iss_closed": 0, "iss_open": 0,
            })
            continue

        # 'issues_found'/'critical' — тоже финальные статусы обхода (обход
        # закончен, просто нашли нарушения), а не промежуточные наравне с
        # 'in_progress' — без них "завершено" занижалось.
        insp_total, insp_done = (await db.execute(text(
            "SELECT COUNT(*), COUNT(*) FILTER (WHERE status IN ('completed', 'issues_found', 'critical')) "
            "FROM inspections WHERE site_id = ANY(:sids)"
        ), {"sids": site_ids})).fetchone()

        iss_total, iss_closed = (await db.execute(text(
            "SELECT COUNT(*), COUNT(*) FILTER (WHERE status = 'closed') "
            "FROM issues WHERE site_id = ANY(:sids)"
        ), {"sids": site_ids})).fetchone()

        work_stats.append({
            "name": d.name, "sites": len(site_ids),
            "insp_total": insp_total, "insp_done": insp_done,
            "iss_total": iss_total, "iss_closed": iss_closed,
            "iss_open": iss_total - iss_closed,
        })
    data["work_stats"] = work_stats
    data["total_sites"] = sum(x["sites"] for x in work_stats)
    data["total_insp"] = sum(x["insp_total"] for x in work_stats)
    data["total_insp_done"] = sum(x["insp_done"] for x in work_stats)
    data["total_iss"] = sum(x["iss_total"] for x in work_stats)
    data["total_iss_closed"] = sum(x["iss_closed"] for x in work_stats)

    # ── Топ категорий нарушений по чек-листу ──
    data["top_categories"] = (await db.execute(text(
        "SELECT COALESCE(ci.category, 'Без категории') AS cat, COUNT(*) AS cnt "
        "FROM checklist_answers ca JOIN checklist_items ci ON ci.id = ca.checklist_item_id "
        "WHERE ca.result = 'defect' GROUP BY cat ORDER BY cnt DESC LIMIT 5"
    ))).fetchall()

    # ── Аномальные названия районов (похоже на опечатку/задвоение при импорте) ──
    data["odd_districts"] = [d.name for d in districts if ";" in d.name or "," in d.name]

    # ── Лидеры дня (обходов начато сегодня, по Москве) ──
    today_msk = datetime.now(MSK).date()
    data["today_msk"] = today_msk
    # Группируем по u.id, не по full_name — у ФИО нет уникальности в базе,
    # инспекторы-тёзки иначе схлопывались бы в одну строку с суммой обходов.
    data["leaders_today"] = (await db.execute(text(
        "SELECT u.id, u.full_name, COUNT(*) AS cnt FROM inspections i "
        "JOIN users u ON u.id = i.inspector_id "
        "WHERE date(i.created_at AT TIME ZONE 'Europe/Moscow') = :today "
        "GROUP BY u.id, u.full_name ORDER BY cnt DESC LIMIT 5"
    ), {"today": today_msk})).fetchall()

    # ── Обходы (полный лог) ──
    inspections = (await db.execute(text(
        "SELECT i.created_at, dist.name AS dist_name, c.name AS court_name, s.type AS site_type, "
        "u.full_name AS inspector, u.phone, i.status, i.started_at, i.completed_at, i.comment, i.id "
        "FROM inspections i "
        "JOIN sites s ON s.id = i.site_id "
        "JOIN courtyards c ON c.id = s.courtyard_id "
        "JOIN districts dist ON dist.id = c.district_id "
        "JOIN users u ON u.id = i.inspector_id "
        "ORDER BY i.created_at DESC"
    ))).fetchall()

    answers_by_insp = {}
    for row in (await db.execute(text(
        "SELECT inspection_id, "
        "COUNT(*) FILTER (WHERE result = 'ok') AS ok_cnt, "
        "COUNT(*) FILTER (WHERE result = 'defect') AS defect_cnt "
        "FROM checklist_answers GROUP BY inspection_id"
    ))).fetchall():
        answers_by_insp[row.inspection_id] = (row.ok_cnt, row.defect_cnt)

    photo_by_insp = {}
    for row in (await db.execute(text(
        "SELECT inspection_id, COUNT(*) AS cnt FROM photos "
        "WHERE target_type IN ('inspection', 'checklist_answer') AND inspection_id IS NOT NULL "
        "GROUP BY inspection_id"
    ))).fetchall():
        photo_by_insp[row.inspection_id] = row.cnt

    insp_rows = []
    for r in inspections:
        ok_cnt, defect_cnt = answers_by_insp.get(r.id, (0, 0))
        insp_rows.append((
            _fmt_dt(r.created_at), r.dist_name, r.court_name, r.site_type,
            r.inspector, r.phone or "", INSPECTION_STATUS_RU.get(r.status, r.status),
            _fmt_dt(r.started_at), _fmt_dt(r.completed_at),
            ok_cnt, defect_cnt, photo_by_insp.get(r.id, 0), r.comment or "",
        ))
    data["insp_rows"] = insp_rows

    # ── Нарушения по чек-листу (детально) ──
    defects = (await db.execute(text(
        "SELECT i.created_at, dist.name AS dist_name, c.name AS court_name, s.type AS site_type, "
        "ci.category, ci.question, ca.comment, u.full_name AS inspector "
        "FROM checklist_answers ca "
        "JOIN checklist_items ci ON ci.id = ca.checklist_item_id "
        "JOIN inspections i ON i.id = ca.inspection_id "
        "JOIN sites s ON s.id = i.site_id "
        "JOIN courtyards c ON c.id = s.courtyard_id "
        "JOIN districts dist ON dist.id = c.district_id "
        "JOIN users u ON u.id = i.inspector_id "
        "WHERE ca.result = 'defect' "
        "ORDER BY dist.name, i.created_at DESC"
    ))).fetchall()
    data["defect_rows"] = [
        (_fmt_dt(r.created_at), r.dist_name, r.court_name, r.site_type,
         r.category or "", r.question, r.comment or "", r.inspector)
        for r in defects
    ]

    # ── Замечания ──
    issues = (await db.execute(text(
        "SELECT iss.created_at, dist.name AS dist_name, c.name AS court_name, iss.title, "
        "iss.description, iss.criticality, iss.status, author.full_name AS author_name, "
        "assignee.full_name AS assignee_name, iss.due_date, iss.fixed_at "
        "FROM issues iss "
        "JOIN sites s ON s.id = iss.site_id "
        "JOIN courtyards c ON c.id = s.courtyard_id "
        "JOIN districts dist ON dist.id = c.district_id "
        "JOIN users author ON author.id = iss.created_by "
        "LEFT JOIN users assignee ON assignee.id = iss.assigned_to "
        "ORDER BY iss.created_at DESC"
    ))).fetchall()
    # Для диаграмм на листе "Замечания" — те же строки, но посчитанные, а
    # не построчно (столбец пустых значений не годится диаграмме)
    data["issues_by_criticality"] = Counter(CRITICALITY_RU.get(r.criticality, r.criticality) for r in issues)
    data["issues_by_status"] = Counter(ISSUE_STATUS_RU.get(r.status, r.status) for r in issues)
    data["issue_rows"] = [
        (_fmt_dt(r.created_at), r.dist_name, r.court_name, r.title, r.description or "",
         CRITICALITY_RU.get(r.criticality, r.criticality), ISSUE_STATUS_RU.get(r.status, r.status),
         r.author_name, r.assignee_name or "", _fmt_dt(r.due_date), _fmt_dt(r.fixed_at))
        for r in issues
    ]

    # ── Просроченные замечания ──
    overdue = (await db.execute(text(
        "SELECT iss.title, dist.name AS dist_name, c.name AS court_name, author.full_name AS author_name, "
        "iss.created_at, iss.due_date, iss.description "
        "FROM issues iss "
        "JOIN sites s ON s.id = iss.site_id "
        "JOIN courtyards c ON c.id = s.courtyard_id "
        "JOIN districts dist ON dist.id = c.district_id "
        "JOIN users author ON author.id = iss.created_by "
        "WHERE iss.status = 'overdue' ORDER BY iss.due_date"
    ))).fetchall()
    data["overdue_rows"] = [
        (r.title, r.dist_name, r.court_name, r.author_name,
         _fmt_dt(r.created_at), _fmt_dt(r.due_date), r.description or "")
        for r in overdue
    ]

    # ── Динамика: число отмеченных пунктов чек-листа по дням на инспектора ──
    # Группируем по u.id — у ФИО нет уникальности, инспекторы-тёзки иначе
    # схлопывались бы в один столбец с суммой на двоих.
    dyn_rows = (await db.execute(text(
        "SELECT date(ca.created_at AT TIME ZONE 'Europe/Moscow') AS d, u.id, u.full_name, COUNT(*) AS cnt "
        "FROM checklist_answers ca "
        "JOIN inspections i ON i.id = ca.inspection_id "
        "JOIN users u ON u.id = i.inspector_id "
        "GROUP BY d, u.id, u.full_name ORDER BY d DESC, u.full_name"
    ))).fetchall()
    day_data: dict = defaultdict(lambda: defaultdict(int))
    inspector_names: dict = {}
    for r in dyn_rows:
        day_data[str(r.d)][r.id] = r.cnt
        inspector_names[r.id] = r.full_name
    sorted_inspector_ids = sorted(inspector_names, key=lambda i: (inspector_names[i], str(i)))
    sorted_inspectors = [inspector_names[i] for i in sorted_inspector_ids]
    data["dynamics_inspectors"] = sorted_inspectors
    data["dynamics_rows"] = [
        tuple([d] + [day_data[d].get(insp_id, 0) for insp_id in sorted_inspector_ids])
        for d in sorted(day_data.keys(), reverse=True)
    ]

    return data


def _sheet(wb, title, headers, rows, widths):
    from openpyxl.utils import get_column_letter
    from app.services.xlsx_style import style_header_row, style_data_row

    ws = wb.create_sheet(title)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for row in rows:
        ws.append(row)
        style_data_row(ws, ws.max_row, len(headers))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def build_workbook(data: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from app.services.xlsx_style import (
        style_header_row, style_data_row, style_header_cell, style_data_cell,
        style_merged_label, CENTER_WRAP, SPACER_ROW_HEIGHT,
    )

    RED = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    BOLD = Font(bold=True)

    def _pie(ws, title, cats_ref, data_ref, anchor):
        chart = PieChart()
        chart.title = title
        chart.height, chart.width = 7, 10
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showPercent = True
        ws.add_chart(chart, anchor)

    def _bar(ws, title, cats_ref, data_ref, anchor, y_title=""):
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.y_axis.title = y_title
        chart.height, chart.width = 9, 18
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, anchor)

    wb = Workbook()

    # ── Обзор ──
    ov = wb.active
    ov.title = "Обзор"
    now_str = datetime.now(MSK).strftime("%d.%m.%Y %H:%M")

    def row(*vals, style=None):
        ov.append(list(vals) if vals else [None])
        r = ov.max_row
        if not vals:
            ov.row_dimensions[r].height = SPACER_ROW_HEIGHT
            return
        if style == "header":
            style_merged_label(ov, r, 2, header=True)
        elif style == "data":
            if len(vals) == 1:
                style_merged_label(ov, r, 2, header=False)
            else:
                for i in range(1, len(vals) + 1):
                    style_data_cell(ov.cell(r, i))

    row("Сводный отчёт по проекту «Журнал обхода площадок» — САО г. Москвы")
    ov["A1"].font = Font(bold=True, size=14)
    ov["A1"].alignment = CENTER_WRAP
    row(f"Снимок на {now_str} (МСК)")
    ov[f"A{ov.max_row}"].alignment = CENTER_WRAP
    row()
    row("Ключевые цифры", style="header")
    row("Регистрация по округу", f"{data['total_reg']} / {data['total_all_invited']} ({data['overall_reg_pct']:.0%})", style="data")
    row("Площадок в системе", data["total_sites"], style="data")
    row("Обходов всего / завершено", f"{data['total_insp']} / {data['total_insp_done']} (в процессе: {data['total_insp'] - data['total_insp_done']})", style="data")
    row("Нарушений создано / открыто сейчас", f"{data['total_iss']} / {data['total_iss'] - data['total_iss_closed']} (закрыто: {data['total_iss_closed']})", style="data")
    row()
    row("Требуют внимания в первую очередь", style="header")

    low_reg = [x for x in data["reg_stats"] if x["pct"] < 0.5 and x["total"] > 0]
    if low_reg:
        names = ", ".join(f"{x['name']} ({x['pct']:.0%})" for x in low_reg)
        row(f"• Регистрация < 50%: {names} — стоит выяснить, не мешает ли что-то войти в систему.", style="data")

    not_registered_admins = data["pending_admins"] + data["pending_okrug_reviewers"]
    if not_registered_admins:
        names = ", ".join(r.full_name for r in not_registered_admins)
        row(f"• Не зарегистрированы админы/окружные проверяющие ({len(not_registered_admins)}): {names} — у них административные права, стоит дожать в первую очередь.", style="data")

    if data["top_categories"]:
        cats = ", ".join(f"{r.cat} ({r.cnt})" for r in data["top_categories"])
        row(f"• Систематические типы нарушений по округу: {cats}.", style="data")

    if data["odd_districts"]:
        row(f"• Похоже на опечатку/задвоение района при заведении в систему: {', '.join(data['odd_districts'])} — не отдельный реальный район.", style="data")

    row()
    row("Лидеры дня по личной активности (обходов начато сегодня)", style="header")
    for r in data["leaders_today"]:
        row(f"   {r.full_name}", r.cnt, style="data")

    row()
    row("Состав отчёта", style="header")
    for name, desc in [
        ("Регистрация", "Кто зарегистрирован/нет по районам, поимённо, худшие районы сверху"),
        ("Сводка по районам", "Площадки/обходы/нарушения + % регистрации, в одной таблице"),
        ("Обходы", "Детальный лог всех обходов: инспектор, площадка, статус, ОК/дефектов/фото"),
        ("Нарушения по чек-листу", "Разбивка нарушений по категориям и конкретным пунктам чек-листа"),
        ("Замечания", "Все зафиксированные замечания с критичностью и статусом"),
        ("Просроченные замечания", "Замечания с истёкшим сроком устранения"),
        ("Динамика", "Активность каждого сотрудника по дням (число отмеченных пунктов чек-листа)"),
    ]:
        row(name, desc, style="data")

    for col, w in zip("ABCDEF", (55, 30, 20, 15, 15, 15)):
        ov.column_dimensions[col].width = w

    # Вспомогательные данные для диаграмм — в стороне от читаемого текста
    # (тот в колонках A-B), чтобы не мешать
    ov["H1"] = "Обходы"
    style_header_cell(ov["H1"], fill=False)
    style_header_cell(ov["I1"], fill=False)
    ov.merge_cells("H1:I1")
    ov["H2"], ov["I2"] = "Завершено", data["total_insp_done"]
    ov["H3"], ov["I3"] = "В процессе", data["total_insp"] - data["total_insp_done"]
    for cell in (ov["H2"], ov["I2"], ov["H3"], ov["I3"]):
        style_data_cell(cell)
    ov["H5"] = "Замечания"
    style_header_cell(ov["H5"], fill=False)
    style_header_cell(ov["I5"], fill=False)
    ov.merge_cells("H5:I5")
    ov["H6"], ov["I6"] = "Открыто", data["total_iss"] - data["total_iss_closed"]
    ov["H7"], ov["I7"] = "Закрыто", data["total_iss_closed"]
    for cell in (ov["H6"], ov["I6"], ov["H7"], ov["I7"]):
        style_data_cell(cell)
    if data["total_insp"] > 0:
        _pie(ov, "Обходы: завершено / в процессе",
             Reference(ov, min_col=8, min_row=2, max_row=3), Reference(ov, min_col=9, min_row=2, max_row=3), "K2")
    if data["total_iss"] > 0:
        _pie(ov, "Замечания: открыто / закрыто",
             Reference(ov, min_col=8, min_row=6, max_row=7), Reference(ov, min_col=9, min_row=6, max_row=7), "K16")

    # ── Регистрация ──
    reg_ws = wb.create_sheet("Регистрация")

    def _reg_spacer():
        reg_ws.append([])
        reg_ws.row_dimensions[reg_ws.max_row].height = SPACER_ROW_HEIGHT

    reg_ws.append(["Регистрация пользователей по районам"])
    reg_ws["A1"].font = Font(bold=True, size=14)
    reg_ws["A1"].alignment = CENTER_WRAP
    reg_ws.append([f"Снимок на {now_str} — user_invites + users, дедуп по ФИО"])
    reg_ws[f"A{reg_ws.max_row}"].alignment = CENTER_WRAP
    _reg_spacer()
    reg_ws.append(["Район", "Зарегистрировано", "Всего приглашено", "% регистрации"])
    style_header_row(reg_ws, reg_ws.max_row, 4)
    for x in data["reg_stats"]:
        reg_ws.append([x["name"], x["registered"], x["total"], x["pct"]])
        style_data_row(reg_ws, reg_ws.max_row, 4)
        fill = RED if x["pct"] < 0.5 else (YELLOW if x["pct"] < 0.8 else GREEN)
        reg_ws.cell(reg_ws.max_row, 4).fill = fill
        reg_ws.cell(reg_ws.max_row, 4).number_format = "0%"
    reg_ws.append(["ИТОГО ПО ОКРУГУ", data["total_reg"], data["total_all_invited"], data["overall_reg_pct"]])
    style_header_row(reg_ws, reg_ws.max_row, 4)
    reg_ws.cell(reg_ws.max_row, 4).number_format = "0%"

    def _reg_header(text):
        reg_ws.append([text])
        style_merged_label(reg_ws, reg_ws.max_row, 2, header=True)

    def _reg_item(text):
        reg_ws.append([text])
        style_merged_label(reg_ws, reg_ws.max_row, 2, header=False)

    def _reg_note(text):
        reg_ws.append([text])
        style_merged_label(reg_ws, reg_ws.max_row, 4, header=False)

    _reg_spacer()
    _reg_header("Не зарегистрированы поимённо (по районам, худшие сверху)")
    for x in data["reg_stats"]:
        if not x["pending"]:
            continue
        _reg_header(x["name"])
        for r in x["pending"]:
            _reg_item(f"   {r.full_name}")

    if data["pending_okrug_reviewers"]:
        _reg_spacer()
        _reg_header("Проверяющие без района (весь округ) — не зарегистрированы")
        for r in data["pending_okrug_reviewers"]:
            _reg_item(f"   {r.full_name}")

    if data["pending_admins"]:
        _reg_spacer()
        _reg_header("Админы — не зарегистрированы")
        for r in data["pending_admins"]:
            _reg_item(f"   {r.full_name}")

    _reg_spacer()
    if data["odd_districts"]:
        _reg_note(f"* {', '.join(data['odd_districts'])} — вероятная опечатка/задвоение при заведении района, не отдельный реальный район.")
    _reg_note("Цвет строки: красный < 50% регистрации, жёлтый 50–80%, зелёный ≥ 80%.")
    for col, w in zip("ABCD", (45, 18, 18, 15)):
        reg_ws.column_dimensions[col].width = w

    if data["reg_stats"]:
        n = len(data["reg_stats"])
        _bar(reg_ws, "% регистрации по районам",
             Reference(reg_ws, min_col=1, min_row=5, max_row=4 + n),
             Reference(reg_ws, min_col=4, min_row=4, max_row=4 + n), "F5", y_title="%")

    # ── Сводка по районам (с % регистрации) ──
    reg_pct_by_name = {x["name"]: x["pct"] for x in data["reg_stats"]}
    sum_ws = wb.create_sheet("Сводка по районам")
    sum_ws.append(["Район", "Площадок", "Обходов", "Заверш.", "Замечаний", "Закрыто", "Открыто сейчас", "% регистрации"])
    style_header_row(sum_ws, 1, 8)
    leader_insp_done = max((x["insp_done"] for x in data["work_stats"]), default=0)
    for x in sorted(data["work_stats"], key=lambda x: x["name"]):
        sum_ws.append([
            x["name"], x["sites"], x["insp_total"], x["insp_done"], x["iss_total"],
            x["iss_closed"], x["iss_open"], reg_pct_by_name.get(x["name"], 0),
        ])
        r = sum_ws.max_row
        style_data_row(sum_ws, r, 8)
        sum_ws.cell(r, 8).number_format = "0%"
        if x["insp_total"] == 0 and x["sites"] > 50:
            for c in range(1, 9):
                sum_ws.cell(r, c).fill = RED
        elif x["insp_done"] == leader_insp_done and leader_insp_done > 0:
            for c in range(1, 9):
                sum_ws.cell(r, c).fill = GREEN
    sum_ws.append(["ИТОГО ПО ОКРУГУ", data["total_sites"], data["total_insp"], data["total_insp_done"],
                    data["total_iss"], data["total_iss_closed"], data["total_iss"] - data["total_iss_closed"],
                    data["overall_reg_pct"]])
    style_header_row(sum_ws, sum_ws.max_row, 8)
    sum_ws.cell(sum_ws.max_row, 8).number_format = "0%"
    sum_ws.append(["Цвет строки: красный — 0 обходов при >50 площадках, зелёный — район-лидер по завершённым обходам."])
    style_merged_label(sum_ws, sum_ws.max_row, 8, header=False)
    for col, w in zip("ABCDEFGH", (24, 12, 10, 10, 12, 10, 14, 14)):
        sum_ws.column_dimensions[col].width = w

    if data["work_stats"]:
        n = len(data["work_stats"])
        _bar(sum_ws, "Обходы по районам: всего / завершено",
             Reference(sum_ws, min_col=1, min_row=2, max_row=1 + n),
             Reference(sum_ws, min_col=3, max_col=4, min_row=1, max_row=1 + n), "J2")

    _sheet(wb, "Обходы",
        ["Дата", "Район", "Двор", "Тип площадки", "Инспектор", "Телефон", "Статус", "Начат", "Завершён", "Пунктов ОК", "Дефектов", "Фото", "Комментарий"],
        data["insp_rows"], [16, 20, 40, 20, 24, 16, 14, 16, 16, 12, 10, 7, 40])
    _sheet(wb, "Нарушения по чек-листу",
        ["Дата обхода", "Район", "Двор", "Тип площадки", "Категория", "Пункт чек-листа", "Комментарий инспектора", "Инспектор"],
        data["defect_rows"], [16, 20, 40, 20, 18, 50, 40, 24])
    iss_ws = _sheet(wb, "Замечания",
        ["Дата", "Район", "Двор", "Заголовок", "Описание", "Критичность", "Статус", "Автор", "Назначено", "Срок", "Устранено"],
        data["issue_rows"], [16, 20, 40, 30, 40, 12, 14, 24, 24, 16, 16])

    # Диаграммы по критичности/статусу — данные для них в стороне от
    # таблицы (та занимает колонки A-K)
    if data["issues_by_criticality"]:
        iss_ws["N1"] = "Критичность"
        style_header_cell(iss_ws["N1"], fill=False)
        style_header_cell(iss_ws["O1"], fill=False)
        iss_ws.merge_cells("N1:O1")
        for i, (label, cnt) in enumerate(data["issues_by_criticality"].most_common(), start=2):
            iss_ws.cell(i, 14, label)
            iss_ws.cell(i, 15, cnt)
            style_data_cell(iss_ws.cell(i, 14))
            style_data_cell(iss_ws.cell(i, 15))
        last = 1 + len(data["issues_by_criticality"])
        _pie(iss_ws, "Замечания по критичности",
             Reference(iss_ws, min_col=14, min_row=2, max_row=last), Reference(iss_ws, min_col=15, min_row=2, max_row=last), "P1")
    if data["issues_by_status"]:
        iss_ws["N17"] = "Статус"
        style_header_cell(iss_ws["N17"], fill=False)
        style_header_cell(iss_ws["O17"], fill=False)
        iss_ws.merge_cells("N17:O17")
        for i, (label, cnt) in enumerate(data["issues_by_status"].most_common(), start=18):
            iss_ws.cell(i, 14, label)
            iss_ws.cell(i, 15, cnt)
            style_data_cell(iss_ws.cell(i, 14))
            style_data_cell(iss_ws.cell(i, 15))
        last = 17 + len(data["issues_by_status"])
        _pie(iss_ws, "Замечания по статусу",
             Reference(iss_ws, min_col=14, min_row=18, max_row=last), Reference(iss_ws, min_col=15, min_row=18, max_row=last), "P17")
    _sheet(wb, "Просроченные замечания",
        ["Заголовок", "Район", "Двор", "Автор", "Создано", "Срок", "Описание"],
        data["overdue_rows"], [30, 20, 40, 24, 16, 16, 40])
    if data["dynamics_rows"]:
        _sheet(wb, "Динамика",
            ["Дата"] + data["dynamics_inspectors"],
            data["dynamics_rows"], [12] + [12] * len(data["dynamics_inspectors"]))

    return wb


async def main():
    engine = create_async_engine(DATABASE_URL, echo=False)
    Session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with Session() as db:
        data = await fetch_all(db)
    await engine.dispose()

    wb = build_workbook(data)
    wb.save(OUT_PATH)
    print(f"OK: {OUT_PATH}")


if __name__ == "__main__":
    asyncio.run(main())
