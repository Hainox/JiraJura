"""Регрессия: Issue.due_date — чистая дата (см. reports.py:_fmt_date),
раньше форматировалась через _fmt_dt ("%d.%m.%Y %H:%M"), из-за чего ячейка
"Срок" в /reports/export.xlsx выглядела как "01.09.2026 00:00" — ложное
время, которого у срока устранения нет."""
import io
import os
import uuid
from datetime import date, timedelta

import psycopg2
import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

SYNC_DB_URL = os.environ["DATABASE_URL"].replace("postgresql+asyncpg://", "postgresql://")


def _exec(sql, params=None):
    conn = psycopg2.connect(SYNC_DB_URL)
    conn.autocommit = True
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or {})
    finally:
        conn.close()


def _find_row(ws, headers: list[str], match_column: str, match_value: str) -> dict | None:
    idx = {h: i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx[match_column]] == match_value:
            return {h: row[i] for h, i in idx.items()}
    return None


@pytest.mark.asyncio
async def test_export_due_date_columns_have_no_time_component(client: AsyncClient, admin_headers):
    r = await client.get("/api/v1/districts/", headers=admin_headers)
    district_id = r.json()[0]["id"]

    me = await client.get("/api/v1/auth/me", headers=admin_headers)
    admin_id = me.json()["id"]

    invite = await client.post("/api/v1/auth/invites", json={
        "login": "DueDateFormatInspector", "full_name": "Инспектор Срок Устранения",
        "role": "inspector", "district_id": district_id,
    }, headers=admin_headers)
    complete = await client.post(
        f"/api/v1/auth/invites/{invite.json()['token']}/complete", json={"password": "DueDate12345"},
    )
    inspector_headers = {"Authorization": f"Bearer {complete.json()['access_token']}"}

    court_id, site_id = str(uuid.uuid4()), str(uuid.uuid4())
    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES (%(c)s, %(d)s, %(n)s)",
        {"c": court_id, "d": district_id, "n": "Двор для проверки срока устранения"},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "(%(s)s, %(c)s, 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((45.0 55.8,45.001 55.8,45.001 55.801,45.0 55.801,45.0 55.8))', 4326), true)",
        {"s": site_id, "c": court_id},
    )
    start = await client.post("/api/v1/inspections/", json={"site_id": site_id, "type": "regular"}, headers=inspector_headers)
    insp_id = start.json()["id"]

    # Открытое замечание со сроком в будущем — попадает в лист "Замечания",
    # но не в "Просроченные замечания" (там due_date < сегодня).
    open_due = date.today() + timedelta(days=5)
    open_issue_id = str(uuid.uuid4())
    _exec(
        "INSERT INTO issues (id, inspection_id, site_id, category_id, title, status, created_by, due_date) VALUES "
        "(%(i)s, %(insp)s, %(s)s, (SELECT id FROM issue_categories WHERE name='Прочее'), "
        "'Открытое замечание со сроком', 'open', %(u)s, %(due)s)",
        {"i": open_issue_id, "insp": insp_id, "s": site_id, "u": admin_id, "due": open_due},
    )

    # Просроченное замечание — due_date в прошлом, статус ещё "открытый".
    overdue_due = date.today() - timedelta(days=3)
    overdue_issue_id = str(uuid.uuid4())
    _exec(
        "INSERT INTO issues (id, inspection_id, site_id, category_id, title, status, created_by, due_date) VALUES "
        "(%(i)s, %(insp)s, %(s)s, (SELECT id FROM issue_categories WHERE name='Прочее'), "
        "'Просроченное замечание', 'open', %(u)s, %(due)s)",
        {"i": overdue_issue_id, "insp": insp_id, "s": site_id, "u": admin_id, "due": overdue_due},
    )

    r = await client.get("/api/v1/reports/export.xlsx", headers=admin_headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))

    expected_open = open_due.strftime("%d.%m.%Y")
    expected_overdue = overdue_due.strftime("%d.%m.%Y")

    iss_ws = wb["Замечания"]
    iss_headers = [c.value for c in iss_ws[1]]
    row = _find_row(iss_ws, iss_headers, "Заголовок", "Открытое замечание со сроком")
    assert row is not None, "Замечание не найдено на листе 'Замечания'"
    assert row["Срок"] == expected_open
    assert ":" not in row["Срок"]

    overdue_ws = wb["Просроченные замечания"]
    overdue_headers = [c.value for c in overdue_ws[1]]
    overdue_row = _find_row(overdue_ws, overdue_headers, "Заголовок", "Просроченное замечание")
    assert overdue_row is not None, "Замечание не найдено на листе 'Просроченные замечания'"
    assert overdue_row["Срок"] == expected_overdue
    assert ":" not in overdue_row["Срок"]
