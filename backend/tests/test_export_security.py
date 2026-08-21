"""Регрессия: текст, введённый пользователями без премодерации (замечания
инспектора, обращения с публичной формы), не должен превращаться в
исполняемую формулу при открытии выгруженного .xlsx в Excel (CWE-1236).
См. app/services/xlsx_style.py:safe_append и его использование в
reports.py/feedback.py. Заодно: /feedback/* — теперь строго admin, не
reviewer (FeedbackReport не привязан к району)."""
import io
import os
import uuid

import psycopg2
import pytest
from openpyxl import load_workbook

SYNC_DB_URL = os.environ["DATABASE_URL"].replace("postgresql+asyncpg://", "postgresql://")

_PAYLOAD = '=HYPERLINK("http://evil.example","click me")'


def _exec(sql, params=None):
    conn = psycopg2.connect(SYNC_DB_URL)
    conn.autocommit = True
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or {})
    finally:
        conn.close()


def _active_category_id() -> str:
    conn = psycopg2.connect(SYNC_DB_URL)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM issue_categories WHERE is_active = true ORDER BY sort_order, name LIMIT 1")
            return str(cur.fetchone()[0])
    finally:
        conn.close()


@pytest.mark.asyncio
async def test_issue_export_neutralizes_formula_injection(client, admin_headers):
    did, court_id, site_id = str(uuid.uuid4()), str(uuid.uuid4()), str(uuid.uuid4())
    _exec(
        "INSERT INTO districts (id, name, code) VALUES (%(id)s, %(name)s, %(code)s)",
        {"id": did, "name": "Формула-инъекция", "code": f"inj_{did[:8]}"},
    )
    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES (%(c)s, %(d)s, %(n)s)",
        {"c": court_id, "d": did, "n": "Двор инъекции"},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "(%(s)s, %(c)s, 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((39.0 55.8,39.001 55.8,39.001 55.801,39.0 55.801,39.0 55.8))', 4326), true)",
        {"s": site_id, "c": court_id},
    )
    start = await client.post("/api/v1/inspections/", json={"site_id": site_id, "type": "regular"}, headers=admin_headers)
    insp_id = start.json()["id"]

    created = await client.post("/api/v1/issues/", json={
        "inspection_id": insp_id, "category_id": _active_category_id(),
        "title": _PAYLOAD, "description": _PAYLOAD, "criticality": "medium",
    }, headers=admin_headers)
    assert created.status_code == 200, created.text

    r = await client.get(
        "/api/v1/reports/export.xlsx", params={"district_id": did}, headers=admin_headers,
    )
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Замечания"]

    found = False
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == _PAYLOAD:
                found = True
                # Ключевая проверка: ячейка сохранена как ТЕКСТ, не как
                # формула (data_type='f' означало бы, что Excel её исполнит).
                assert cell.data_type == "s", (
                    f"payload written as formula (data_type={cell.data_type!r}), "
                    "not neutralized"
                )
    assert found, "payload not found in export at all"


@pytest.mark.asyncio
async def test_feedback_export_neutralizes_formula_injection(client, admin_headers):
    submit = await client.post("/api/v1/feedback/", json={
        "report_type": "app", "message": _PAYLOAD,
    })
    assert submit.status_code == 201, submit.text

    r = await client.get("/api/v1/feedback/export.xlsx", headers=admin_headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Обращения"]

    found = False
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == _PAYLOAD:
                found = True
                assert cell.data_type == "s"
    assert found


@pytest.mark.asyncio
async def test_feedback_endpoints_reject_reviewer(client, admin_headers):
    """FeedbackReport не привязан к району — reviewer видел бы обращения
    ВСЕГО округа, шире, чем где-либо ещё. Держим строго admin."""
    r = await client.get("/api/v1/districts/", headers=admin_headers)
    district_id = r.json()[0]["id"]

    invite = await client.post("/api/v1/auth/invites", json={
        "login": "FeedbackReviewer", "full_name": "Обращенский Проверяющий",
        "role": "reviewer", "district_id": district_id,
    }, headers=admin_headers)
    complete = await client.post(
        f"/api/v1/auth/invites/{invite.json()['token']}/complete", json={"password": "Feedback12345"},
    )
    reviewer_headers = {"Authorization": f"Bearer {complete.json()['access_token']}"}

    listed = await client.get("/api/v1/feedback/", headers=reviewer_headers)
    assert listed.status_code == 403, listed.text

    exported = await client.get("/api/v1/feedback/export.xlsx", headers=reviewer_headers)
    assert exported.status_code == 403, exported.text
