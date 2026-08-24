"""Reports router."""
import io
from datetime import date, datetime, timezone
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.database import get_db
from app.models import (
    District, Courtyard, Site, Inspection, Issue, ChecklistAnswer, ChecklistItem, User, IssueStatusHistory,
)
from app.schemas import ReportWeeklyOut, ReportMonthlyOut, DashboardDistrictRow, DashboardOut
from app.services.permissions import require_role
from app.services.timezone import MSK, msk_day_bounds_utc
from app.services.statistics import StatisticsService
from app.services.statistics.filters import build_filter

router = APIRouter()

# Обход считается завершённым, даже если по итогу нашлись нарушения —
# 'issues_found'/'critical' это тоже финальные статусы (проставляются вместе
# с completed_at, см. update_inspection в inspections.py), а не промежуточные
# наравне с 'in_progress'. Без этого списка "Завершено" в дашборде/отчётах
# занижался, а "В процессе" — по факту не завышался, но обходы с найденными
# нарушениями просто выпадали из обеих колонок.
INSPECTION_DONE_STATUSES = ("completed", "issues_found", "critical")


@router.get("/weekly", response_model=list[ReportWeeklyOut])
async def weekly_report(
    district_id: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("reviewer", "admin")),
):
    did = UUID(district_id) if district_id else None
    dashboard = await StatisticsService(
        db, build_filter(current_user, date_from, date_to, did)
    ).dashboard()
    return [ReportWeeklyOut(
        district_id=row.district_id, district_name=row.district_name,
        total_sites=row.total_sites, inspected_sites=row.sites_inspected,
        issues_open=row.issues_open + row.issues_in_work,
        issues_overdue=row.issues_overdue,
    ) for row in dashboard.districts]


@router.get("/monthly", response_model=list[ReportMonthlyOut])
async def monthly_report(
    district_id: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("reviewer", "admin")),
):
    today = datetime.now(MSK).date()
    month_start = today.replace(day=1)
    did = UUID(district_id) if district_id else None
    filters = build_filter(
        current_user, date_from or month_start, date_to or today, did
    )
    dashboard = await StatisticsService(db, filters).dashboard()
    closure_stmt = (
        select(Courtyard.district_id, func.count(IssueStatusHistory.id))
        .join(Issue, Issue.id == IssueStatusHistory.issue_id)
        .join(Site, Site.id == Issue.site_id)
        .join(Courtyard, Courtyard.id == Site.courtyard_id)
        .where(IssueStatusHistory.new_status == "closed",
               IssueStatusHistory.created_at >= filters.start_utc,
               IssueStatusHistory.created_at < filters.end_utc)
        .group_by(Courtyard.district_id)
    )
    if filters.district_id:
        closure_stmt = closure_stmt.where(Courtyard.district_id == filters.district_id)
    closures = {str(key): value for key, value in (await db.execute(closure_stmt)).all()}
    return [ReportMonthlyOut(
        district_id=row.district_id, district_name=row.district_name,
        total_sites=row.total_sites, inspected_sites=row.sites_inspected,
        issues_created=row.issues_found,
        issues_closed=int(closures.get(row.district_id, 0)),
        issues_overdue=row.issues_overdue,
    ) for row in dashboard.districts]


# ── Дашборд админа ─────────────────────────────────────────────

@router.get("/dashboard", response_model=DashboardOut)
async def admin_dashboard(
    district_id: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("reviewer", "admin")),
):
    """Сводка по районам для дашборда: обходы, замечания, статусы."""
    if current_user.role == "reviewer" and current_user.district_id is not None:
        district_id = str(current_user.district_id)

    dt_from, dt_to = msk_day_bounds_utc(date_from, date_to)

    def period(q, column):
        if dt_from is not None:
            q = q.where(column >= dt_from)
        if dt_to is not None:
            q = q.where(column < dt_to)
        return q

    dist_q = select(District).order_by(District.name)
    if district_id:
        dist_q = dist_q.where(District.id == district_id)
    districts = (await db.execute(dist_q)).scalars().all()

    rows: list[DashboardDistrictRow] = []
    total_row = DashboardDistrictRow(
        district_id=UUID("00000000-0000-0000-0000-000000000000"),
        district_name="ВСЕГО", total_sites=0,
        sites_inspected=0, sites_not_inspected=0,
        inspections_total=0, inspections_completed=0, inspections_in_progress=0,
        inspections_ok=0, inspections_with_defects=0,
        checklist_defects=0,
        issues_total=0, issues_open=0, issues_fixed=0,
        issues_revision_needed=0, issues_closed=0, issues_overdue=0,
    )

    for d in districts:
        site_sub = (
            select(Site.id)
            .join(Courtyard, Site.courtyard_id == Courtyard.id)
            .where(Courtyard.district_id == d.id)
        )

        total_sites = (await db.execute(
            select(func.count()).select_from(Site).where(Site.id.in_(site_sub), Site.is_active)
        )).scalar_one() or 0

        insp_base = select(func.count()).select_from(Inspection).where(Inspection.site_id.in_(site_sub))
        inspections_total = (await db.execute(period(insp_base, Inspection.created_at))).scalar_one() or 0
        inspections_completed = (await db.execute(period(
            select(func.count()).select_from(Inspection).where(Inspection.site_id.in_(site_sub), Inspection.status.in_(INSPECTION_DONE_STATUSES)),
            Inspection.created_at,
        ))).scalar_one() or 0
        inspections_in_progress = (await db.execute(period(
            select(func.count()).select_from(Inspection).where(Inspection.site_id.in_(site_sub), Inspection.status == "in_progress"),
            Inspection.created_at,
        ))).scalar_one() or 0

        # Охват: сколько РАЗНЫХ площадок доведено до финального статуса за
        # период, а не сколько обходов записано. Это «проверено площадок» —
        # сюда идут и полностью «зелёные» обходы (0 дефектов), которые иначе
        # нигде не отражались бы, и район с идеальным порядком не выглядел бы
        # так же, как район, где вообще не обходили.
        sites_inspected = (await db.execute(period(
            select(func.count(func.distinct(Inspection.site_id)))
            .select_from(Inspection)
            .where(Inspection.site_id.in_(site_sub), Inspection.status.in_(INSPECTION_DONE_STATUSES)),
            Inspection.created_at,
        ))).scalar_one() or 0
        sites_not_inspected = max(total_sites - sites_inspected, 0)

        # Результат завершённых обходов: «с нарушениями» — если в обходе есть
        # хотя бы один defect-пункт чек-листа; «без» — все остальные
        # завершённые. Дефект виден сразу в момент обхода, не дожидаясь, пока
        # кто-то прикрепит фото исправления и закроет замечание (раньше для
        # таких обходов «что-то менялось» только после закрытия issue).
        inspections_with_defects = (await db.execute(period(
            select(func.count(func.distinct(Inspection.id)))
            .select_from(Inspection)
            .join(ChecklistAnswer, ChecklistAnswer.inspection_id == Inspection.id)
            .where(
                Inspection.site_id.in_(site_sub),
                Inspection.status.in_(INSPECTION_DONE_STATUSES),
                ChecklistAnswer.result == "defect",
            ),
            Inspection.created_at,
        ))).scalar_one() or 0
        inspections_ok = max(inspections_completed - inspections_with_defects, 0)

        # Реально выявлено при обходе — по чек-листу, независимо от того,
        # оформил ли инспектор отдельное замечание (см. комментарий у поля
        # в schemas.py). Именно это число надо использовать для оценки
        # района, а не issues_total ниже.
        defects_base = (
            select(func.count()).select_from(ChecklistAnswer)
            .join(Inspection, ChecklistAnswer.inspection_id == Inspection.id)
            .where(Inspection.site_id.in_(site_sub), ChecklistAnswer.result == "defect")
        )
        checklist_defects = (await db.execute(period(defects_base, ChecklistAnswer.created_at))).scalar_one() or 0

        iss_base = select(func.count()).select_from(Issue).where(Issue.site_id.in_(site_sub))
        issues_total = (await db.execute(period(iss_base, Issue.created_at))).scalar_one() or 0
        # issues_open/fixed/revision_needed/closed/overdue — ТЕКУЩИЙ статус
        # (снимок "сейчас"), не событие внутри периода: замечание, оформленное
        # до начала выбранного диапазона дат, но до сих пор не устранённое,
        # обязано попасть в issues_open/issues_revision_needed — иначе оно
        # необъяснимо "пропадает" из дашборда, хотя объективно требует
        # внимания прямо сейчас. Период фильтрует только issues_total
        # (сколько ОФОРМЛЕНО за период — это уже реальное событие с датой).
        issues_open = (await db.execute(
            select(func.count()).select_from(Issue).where(
                Issue.site_id.in_(site_sub), Issue.status.in_(["open", "assigned", "in_work"])
            ),
        )).scalar_one() or 0
        # «fixed» + «control»: исправлено и ждёт приёмки / на контроле —
        # обе стадии означают «устранено, но ещё не принято», раньше «control»
        # выпадал из всех вёдер дашборда и такие замечания «терялись».
        issues_fixed = (await db.execute(
            select(func.count()).select_from(Issue).where(Issue.site_id.in_(site_sub), Issue.status.in_(["fixed", "control"])),
        )).scalar_one() or 0
        issues_revision_needed = (await db.execute(
            select(func.count()).select_from(Issue).where(Issue.site_id.in_(site_sub), Issue.status == "revision_needed"),
        )).scalar_one() or 0
        issues_closed = (await db.execute(
            select(func.count()).select_from(Issue).where(Issue.site_id.in_(site_sub), Issue.status == "closed"),
        )).scalar_one() or 0
        issues_overdue = (await db.execute(
            select(func.count()).select_from(Issue).where(
                Issue.site_id.in_(site_sub),
                Issue.status.in_(("open", "assigned", "in_work", "revision_needed")),
                Issue.due_date.is_not(None),
                Issue.due_date < datetime.now(MSK).date(),
            ),
        )).scalar_one() or 0

        row = DashboardDistrictRow(
            district_id=d.id, district_name=d.name, total_sites=total_sites,
            sites_inspected=sites_inspected, sites_not_inspected=sites_not_inspected,
            inspections_total=inspections_total, inspections_completed=inspections_completed,
            inspections_in_progress=inspections_in_progress,
            inspections_ok=inspections_ok, inspections_with_defects=inspections_with_defects,
            checklist_defects=checklist_defects,
            issues_total=issues_total, issues_open=issues_open,
            issues_fixed=issues_fixed, issues_revision_needed=issues_revision_needed,
            issues_closed=issues_closed, issues_overdue=issues_overdue,
        )
        rows.append(row)

        # суммируем в totals
        total_row.total_sites += total_sites
        total_row.sites_inspected += sites_inspected
        total_row.sites_not_inspected += sites_not_inspected
        total_row.inspections_total += inspections_total
        total_row.inspections_completed += inspections_completed
        total_row.inspections_in_progress += inspections_in_progress
        total_row.inspections_ok += inspections_ok
        total_row.inspections_with_defects += inspections_with_defects
        total_row.checklist_defects += checklist_defects
        total_row.issues_total += issues_total
        total_row.issues_open += issues_open
        total_row.issues_fixed += issues_fixed
        total_row.issues_revision_needed += issues_revision_needed
        total_row.issues_closed += issues_closed
        total_row.issues_overdue += issues_overdue

    return DashboardOut(districts=rows, totals=total_row)


# ── Выгрузка в Excel ─────────────────────────────────────────────

INSPECTION_STATUS_RU = {
    "planned": "Запланирован", "in_progress": "В процессе",
    "completed": "Завершён", "issues_found": "Есть нарушения",
    "critical": "Критический",
}
ISSUE_STATUS_RU = {
    "open": "Открыто", "assigned": "Назначено", "in_work": "В работе",
    "fixed": "Устранено", "control": "На контроле", "closed": "Закрыто",
    "overdue": "Просрочено", "revision_needed": "На доработке",
}
CRITICALITY_RU = {
    "low": "Низкая", "medium": "Средняя", "high": "Высокая",
    "critical": "Критическая",
}


def _fmt_dt(value) -> str:
    return value.strftime("%d.%m.%Y %H:%M") if value else ""


def _sheet(wb, title: str, headers: list[str], rows: list[tuple], widths: list[int]):
    from openpyxl.utils import get_column_letter
    from app.services.xlsx_style import style_header_row, style_data_row, safe_append

    ws = wb.create_sheet(title)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for row in rows:
        safe_append(ws, row)
        style_data_row(ws, ws.max_row, len(headers))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


@router.get("/export.xlsx")
async def export_xlsx(
    district_id: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    all_time: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("reviewer", "admin")),
):
    """Выгрузка журнала в Excel: сводка по районам, обходы, замечания.

    reviewer с заданным district_id видит только свой район — параметр
    district_id для него принудительно замещается собственным районом.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint
    from app.services.xlsx_style import (
        style_header_cell, style_data_cell,
        style_merged_label, safe_append, CENTER_WRAP, SPACER_ROW_HEIGHT,
    )

    if current_user.role == "reviewer" and current_user.district_id is not None:
        district_id = str(current_user.district_id)

    effective_district = UUID(district_id) if district_id else None
    stats_filter = build_filter(current_user, date_from, date_to, effective_district, all_time=all_time)
    # Сохраняем legacy-контракт детальных листов: отсутствие дат означает
    # всю историю. Районная сводка ниже использует календарный default v2.
    export_from = stats_filter.date_from if all_time else date_from
    export_to = stats_filter.date_to if all_time else date_to
    dt_from, dt_to = msk_day_bounds_utc(export_from, export_to)

    def period(q, column):
        if dt_from is not None:
            q = q.where(column >= dt_from)
        if dt_to is not None:
            q = q.where(column < dt_to)
        return q

    # ── Топ районов по устранению замечаний — закрытие замечания это
    # СОБЫТИЕ (переход статуса в 'closed'), а не снимок текущего состояния,
    # поэтому считаем по IssueStatusHistory с фильтром периода, а не по
    # Issue.status: иначе "закрыто вчера и потом переоткрыто" неотличимо от
    # "закрыто только что", а сегодняшнее закрытие замечания, оформленного
    # неделю назад, не попало бы в подсчёт вовсе. ──
    closures_q = (
        select(District.id, District.name, func.count())
        .select_from(IssueStatusHistory)
        .join(Issue, IssueStatusHistory.issue_id == Issue.id)
        .join(Site, Issue.site_id == Site.id)
        .join(Courtyard, Site.courtyard_id == Courtyard.id)
        .join(District, Courtyard.district_id == District.id)
        .where(IssueStatusHistory.new_status == "closed")
    )
    if district_id:
        closures_q = closures_q.where(District.id == district_id)
    closures_q = period(closures_q, IssueStatusHistory.created_at)
    closures_q = closures_q.group_by(District.id, District.name).order_by(func.count().desc())
    district_closures = (await db.execute(closures_q)).all()
    total_closures_period = sum(cnt for _did, _name, cnt in district_closures)

    # ── Топ категорий нарушений и лидеры дня — та же зона видимости
    # (district_id), что и у остальных листов этой выгрузки ──
    # cat_expr переиспользуется как единый объект и в SELECT, и в GROUP BY —
    # если вместо этого написать func.coalesce(...) дважды, SQLAlchemy
    # биндит каждое вхождение отдельным параметром, и Postgres не признаёт
    # их одним и тем же выражением (GroupingError: must appear in GROUP BY).
    cat_expr = func.coalesce(ChecklistItem.category, "Без категории").label("cat")
    top_cat_q = (
        select(cat_expr, func.count())
        .select_from(ChecklistAnswer)
        .join(ChecklistItem, ChecklistAnswer.checklist_item_id == ChecklistItem.id)
        .where(ChecklistAnswer.result == "defect")
    )
    if district_id:
        top_cat_q = (
            top_cat_q.join(Inspection, ChecklistAnswer.inspection_id == Inspection.id)
            .join(Site, Inspection.site_id == Site.id).join(Courtyard, Site.courtyard_id == Courtyard.id)
            .where(Courtyard.district_id == district_id)
        )
    top_cat_q = top_cat_q.group_by(cat_expr).order_by(func.count().desc()).limit(5)
    top_categories = (await db.execute(top_cat_q)).all()

    odd_districts = [d.name for d in (await db.execute(select(District))).scalars().all() if ";" in d.name or "," in d.name]

    today_msk = datetime.now(MSK).date()
    leaders_q = (
        select(User.id, User.full_name, func.count())
        .select_from(Inspection).join(User, Inspection.inspector_id == User.id)
        .where(func.date(func.timezone("Europe/Moscow", Inspection.created_at)) == today_msk)
    )
    if district_id:
        leaders_q = leaders_q.join(Site, Inspection.site_id == Site.id).join(Courtyard, Site.courtyard_id == Courtyard.id).where(Courtyard.district_id == district_id)
    leaders_q = leaders_q.group_by(User.id, User.full_name).order_by(func.count().desc()).limit(5)
    leaders_today = (await db.execute(leaders_q)).all()

    # ── Обходы ──
    insp_q = (
        select(Inspection, District.name, Courtyard.name, Site.type, User.full_name, User.phone)
        .join(Site, Inspection.site_id == Site.id)
        .join(Courtyard, Site.courtyard_id == Courtyard.id)
        .join(District, Courtyard.district_id == District.id)
        .join(User, Inspection.inspector_id == User.id)
        .order_by(Inspection.created_at.desc())
    )
    if district_id:
        insp_q = insp_q.where(District.id == district_id)
    insp_q = period(insp_q, Inspection.created_at)
    inspections = (await db.execute(insp_q)).all()

    ans_rows = (await db.execute(
        select(
            ChecklistAnswer.inspection_id,
            func.count().filter(ChecklistAnswer.result == "ok"),
            func.count().filter(ChecklistAnswer.result == "defect"),
        ).group_by(ChecklistAnswer.inspection_id)
    )).all()
    answers = {row[0]: (row[1], row[2]) for row in ans_rows}

    # Считаем фото на каждый обход
    from app.models import Photo
    photo_counts = {}
    photo_rows = (await db.execute(
        select(Photo.inspection_id, func.count())
        .where(Photo.target_type.in_(["inspection", "checklist_answer"]))
        .group_by(Photo.inspection_id)
    )).all()
    for insp_id, cnt in photo_rows:
        photo_counts[str(insp_id)] = cnt

    insp_data = []
    for insp, dist_name, court_name, site_type, inspector_name, phone in inspections:
        ok_cnt, defect_cnt = answers.get(insp.id, (0, 0))
        insp_data.append((
            _fmt_dt(insp.created_at), dist_name, court_name, site_type,
            inspector_name, phone or "",
            INSPECTION_STATUS_RU.get(insp.status, insp.status),
            _fmt_dt(insp.started_at), _fmt_dt(insp.completed_at),
            ok_cnt, defect_cnt, photo_counts.get(str(insp.id), 0),
            insp.comment or "",
        ))

    # ── Задания (снимок "сейчас", а не журнал за период) ──
    # Проверяющие просили именно это отдельно от "Обходы" выше: та вкладка —
    # хронологический лог за выбранный период, а тут — по каждой активной
    # площадке в зоне видимости: кто за неё отвечает и когда там были в
    # последний раз (независимо от периода отчёта), чтобы сразу видеть, где
    # давно не было обхода или он вообще не назначен. district_id/scope тот
    # же, что и у остальных листов этой выгрузки — период (dt_from/dt_to)
    # тут намеренно не применяется, это не история, а текущее состояние.
    assign_q = (
        select(Site, Courtyard.name, District.name, User.full_name, User.phone)
        .join(Courtyard, Site.courtyard_id == Courtyard.id)
        .join(District, Courtyard.district_id == District.id)
        .outerjoin(User, Site.assigned_inspector_id == User.id)
        .where(Site.is_active)
    )
    if district_id:
        assign_q = assign_q.where(District.id == district_id)
    site_rows = (await db.execute(assign_q)).all()

    site_ids = [site.id for site, *_ in site_rows]
    last_by_site: dict = {}
    if site_ids:
        last_sub = (
            select(Inspection.site_id, func.max(Inspection.created_at).label("last_dt"))
            .where(Inspection.site_id.in_(site_ids))
            .group_by(Inspection.site_id)
            .subquery()
        )
        last_rows = (await db.execute(
            select(Inspection.site_id, Inspection.created_at, Inspection.status)
            .join(last_sub, (Inspection.site_id == last_sub.c.site_id) & (Inspection.created_at == last_sub.c.last_dt))
        )).all()
        last_by_site = {row.site_id: (row.created_at, row.status) for row in last_rows}

    assignment_rows = []
    for site, court_name, dist_name, insp_name, insp_phone in site_rows:
        last = last_by_site.get(site.id)
        assignment_rows.append((
            last[0] if last else None,  # ключ сортировки, вырежем перед записью в лист
            dist_name, court_name, site.type,
            insp_name or "Не назначена", insp_phone or "",
            _fmt_dt(last[0]) if last else "Обхода ещё не было",
            INSPECTION_STATUS_RU.get(last[1], last[1]) if last else "",
        ))
    # Сначала то, что реально требует внимания: без единого обхода — выше
    # всего, дальше по возрастанию давности последнего визита.
    assignment_rows.sort(key=lambda r: (r[0] is not None, r[0]))
    assignment_data = [row[1:] for row in assignment_rows]

    # ── Замечания ──
    Assignee = aliased(User)
    iss_q = (
        select(Issue, District.name, Courtyard.name, User.full_name, Assignee.full_name)
        .join(Site, Issue.site_id == Site.id)
        .join(Courtyard, Site.courtyard_id == Courtyard.id)
        .join(District, Courtyard.district_id == District.id)
        .join(User, Issue.created_by == User.id)
        .outerjoin(Assignee, Issue.assigned_to == Assignee.id)
        .order_by(Issue.created_at.desc())
    )
    if district_id:
        iss_q = iss_q.where(District.id == district_id)
    iss_q = period(iss_q, Issue.created_at)
    issues = (await db.execute(iss_q)).all()

    iss_data = [
        (
            _fmt_dt(iss.created_at), dist_name, court_name, iss.title,
            iss.description or "", CRITICALITY_RU.get(iss.criticality, iss.criticality),
            ISSUE_STATUS_RU.get(iss.status, iss.status), author_name,
            assignee_name or "", _fmt_dt(iss.due_date), _fmt_dt(iss.fixed_at),
        )
        for iss, dist_name, court_name, author_name, assignee_name in issues
    ]

    # ── Нарушения по чек-листу (детально) ──
    defect_q = (
        select(
            ChecklistAnswer, ChecklistItem.question, ChecklistItem.category,
            Inspection.created_at, District.name, Courtyard.name,
            Site.type, User.full_name,
        )
        .join(Inspection, ChecklistAnswer.inspection_id == Inspection.id)
        .join(ChecklistItem, ChecklistAnswer.checklist_item_id == ChecklistItem.id)
        .join(Site, Inspection.site_id == Site.id)
        .join(Courtyard, Site.courtyard_id == Courtyard.id)
        .join(District, Courtyard.district_id == District.id)
        .join(User, Inspection.inspector_id == User.id)
        .where(ChecklistAnswer.result == "defect")
        .order_by(District.name, Inspection.created_at.desc())
    )
    if district_id:
        defect_q = defect_q.where(District.id == district_id)
    defect_q = period(defect_q, Inspection.created_at)
    defects = (await db.execute(defect_q)).all()

    defect_data = [
        (
            _fmt_dt(insp_dt), dist_name, court_name, site_type,
            category or "", question, answer.comment or "",
            inspector_name,
        )
        for answer, question, category, insp_dt, dist_name, court_name, site_type, inspector_name in defects
    ]

    # ── Сводка по районам: единый statistics v2 service ──
    stats_dashboard = await StatisticsService(db, stats_filter).dashboard()
    summary_data = [
        (
            row.district_name,
            row.total_sites,
            row.sites_inspected,
            row.coverage_pct,
            row.inspections_total,
            row.inspections_green,
            row.inspections_with_defects,
            row.issues_found,
            row.issues_fixed_events,
            row.issues_closed_events,
            row.issues_revision,
            row.issues_not_fixed,
            row.issues_overdue,
            row.issues_closed_pct,
        )
        for row in stats_dashboard.districts
    ]
    totals = stats_dashboard.totals
    summary_rows = [*summary_data, (
        "ИТОГО", totals.total_sites, totals.sites_inspected, totals.coverage_pct,
        totals.inspections_total, totals.inspections_green, totals.inspections_with_defects,
        totals.issues_found, totals.issues_fixed_events, totals.issues_closed_events, totals.issues_revision,
        totals.issues_not_fixed, totals.issues_overdue, totals.issues_closed_pct,
    )]

    # ── Динамика по дням (инспектор × дата) ──
    # Группируем по User.id, а не по ФИО — у full_name нет уникальности в
    # базе, и два инспектора-тёзки схлопывались бы в один столбец с суммой
    # их обходов на двоих.
    # МСК-дата, не func.date() от сырого UTC (как в "Лидеры дня" выше) —
    # иначе обходы с 00:00 до 02:59 МСК (21:00-23:59 UTC предыдущих суток)
    # попадали бы во "вчера", хотя по факту это уже следующий рабочий день.
    from collections import defaultdict
    day_msk = func.date(func.timezone("Europe/Moscow", Inspection.created_at))
    day_stats_q = (
        select(
            day_msk, User.id, User.full_name, func.count()
        )
        .join(User, Inspection.inspector_id == User.id)
        .group_by(day_msk, User.id, User.full_name)
        .order_by(day_msk.desc(), User.full_name)
    )
    if district_id:
        # Явные условия join — БЕЗ них SQLAlchemy между Site и уже
        # присоединённым User выбирает ПЕРВУЮ попавшуюся FK-связь
        # (users.id = sites.assigned_inspector_id — "назначенный
        # инспектор"), а не ту, что реально нужна (inspections.site_id =
        # sites.id). С неявным .join(Site) фильтр по району на этом листе
        # молча резолвился не в те строки и всегда возвращал пусто —
        # найдено этим самым регресс-тестом на МСК-дату.
        day_stats_q = (
            day_stats_q.join(Site, Inspection.site_id == Site.id)
            .join(Courtyard, Site.courtyard_id == Courtyard.id)
            .where(Courtyard.district_id == district_id)
        )
    day_stats_q = period(day_stats_q, Inspection.created_at)
    day_stats = (await db.execute(day_stats_q)).all()

    # группируем: дата → {id инспектора: кол-во}
    day_data: dict[str, dict] = defaultdict(lambda: defaultdict(int))
    inspector_names: dict = {}
    for dt, insp_id, name, cnt in day_stats:
        day_data[str(dt)][insp_id] = cnt
        inspector_names[insp_id] = name
    sorted_inspector_ids = sorted(inspector_names, key=lambda i: (inspector_names[i], str(i)))
    sorted_inspectors = [inspector_names[i] for i in sorted_inspector_ids]
    dynamics_rows = []
    for dt in sorted(day_data.keys(), reverse=True):
        row = [dt] + [day_data[dt].get(insp_id, 0) for insp_id in sorted_inspector_ids]
        dynamics_rows.append(tuple(row))

    # ── Просроченные замечания ──
    overdue_q = (
        select(Issue, District.name, Courtyard.name, User.full_name)
        .join(Site, Issue.site_id == Site.id)
        .join(Courtyard, Site.courtyard_id == Courtyard.id)
        .join(District, Courtyard.district_id == District.id)
        .join(User, Issue.created_by == User.id)
        .where(
            Issue.status.in_(("open", "assigned", "in_work", "revision_needed")),
            Issue.due_date.is_not(None),
            Issue.due_date < datetime.now(MSK).date(),
        )
        .order_by(Issue.due_date)
    )
    if district_id:
        overdue_q = overdue_q.where(District.id == district_id)
    overdue_items = (await db.execute(overdue_q)).all()
    overdue_rows = [
        (iss.title, dist_name, court_name, author_name,
         _fmt_dt(iss.created_at), _fmt_dt(iss.due_date), iss.description or "")
        for iss, dist_name, court_name, author_name in overdue_items
    ]

    # ── Сборка файла ──
    wb = Workbook()

    # Единая палитра для всех диаграмм отчёта — вместо цветов по умолчанию,
    # которые Excel назначает сериям сам (обычно случайный порядок радуги,
    # не связанный со смыслом "хорошо/плохо"). Статусные цвета (good/warning/
    # critical) — для пар "хорошее/плохое" состояние (проверено/не проверено,
    # без нарушений/с нарушениями), категориальный синий — для нейтральных
    # сравнений величин (районы, категории нарушений, обходы по дням).
    CHART_CATEGORICAL = "2A78D6"
    CHART_GOOD = "0CA30C"
    CHART_WARNING = "FAB219"
    CHART_CRITICAL = "D03B3B"
    CHART_MUTED = "898781"

    def _pie(ws, title, cats_ref, data_ref, anchor, colors=None):
        """colors — (hex, hex) для двух долей пирога, по умолчанию без
        переопределения (цвета темы Excel)."""
        chart = PieChart()
        chart.title = title
        chart.height, chart.width = 7, 10
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showPercent = True
        if colors:
            points = []
            for i, color in enumerate(colors):
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = color
                points.append(pt)
            chart.series[0].data_points = points
        ws.add_chart(chart, anchor)

    def _bar(ws, title, cats_ref, data_ref, anchor, y_title="", color=CHART_CATEGORICAL):
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.y_axis.title = y_title
        # openpyxl создаёt обе оси с axPos="l" по умолчанию — для
        # горизонтальной оси категорий колоночного графика это неверно и
        # конфликтует с осью значений (тоже "l"): Excel не может понять,
        # где рисовать подписи категорий, и молча не рисует их вообще.
        chart.x_axis.axPos = "b"
        chart.height, chart.width = 9, 18
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        for s in chart.series:
            s.graphicalProperties.solidFill = color
        ws.add_chart(chart, anchor)

    def _line(ws, title, cats_ref, data_ref, anchor, y_title="", color=CHART_CATEGORICAL):
        chart = LineChart()
        chart.title = title
        chart.y_axis.title = y_title
        chart.x_axis.axPos = "b"  # см. комментарий в _bar — тот же дефолт-баг openpyxl
        chart.height, chart.width = 9, 18
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        for s in chart.series:
            s.graphicalProperties.line.solidFill = color
            s.graphicalProperties.line.width = 20000  # EMU, ≈2pt — тонкая, но заметная линия
            s.smooth = False
            s.marker.symbol = "circle"
            s.marker.size = 5
            s.marker.graphicalProperties.solidFill = color
            s.marker.graphicalProperties.line.solidFill = color
        ws.add_chart(chart, anchor)

    # ── Обзор — тот же лист, что в generate_summary_report.py (раньше
    # доступен только через ручной запуск скрипта на сервере) ──
    now_str = datetime.now(MSK).strftime("%d.%m.%Y %H:%M")
    # Индексы кортежей summary_data (0-based):
    # 0 name, 1 total_sites, 2 sites_inspected, 3 coverage_pct,
    # 4 inspections, 5 without_defects, 6 with_defects, 7 found,
    # 8 fixed_events, 9 closed_events, 10 revision, 11 not_fixed,
    # 12 overdue, 13 closed_pct
    def _sum(idx):
        return sum(x[idx] for x in summary_data) if summary_data else 0

    total_sites_all = _sum(1)
    sites_inspected_all = _sum(2)
    sites_not_inspected_all = max(total_sites_all - sites_inspected_all, 0)
    total_insp_all = _sum(4)
    inspections_ok_all = _sum(5)
    inspections_with_defects_all = _sum(6)
    total_checklist_defects_all = total_iss_all = _sum(7)
    total_iss_open_all = stats_dashboard.totals.issues_open + stats_dashboard.totals.issues_in_work
    total_iss_fixed_all = stats_dashboard.totals.issues_on_check
    total_iss_revision_all = _sum(10)
    total_iss_closed_all = stats_dashboard.totals.issues_closed
    total_iss_overdue_all = _sum(12)
    total_insp_done_all = inspections_ok_all + inspections_with_defects_all
    total_insp_in_progress_all = total_insp_all - total_insp_done_all

    ov = wb.active
    ov.title = "Обзор"

    def _row(*vals, style=None):
        """style="header" — жирный+жёлтый+рамка (раздел-заголовок), сливает
        A:B в один заполненный блок (как в эталоне).
        style="data" — рамка+центр; при одном значении тоже сливает A:B
        (сноска/пункт списка, чтобы не торчать узкой рамкой рядом с пустой
        нестилизованной ячейкой), при двух — рамка на каждой без слияния
        (во второй ячейке реальные данные).
        style=None — как раньше, без оформления (заголовок отчёта и т.п.).
        Пустой вызов _row() — разделитель между секциями, невысокая
        строка, чтобы не растягивать отчёт пустыми промежутками
        стандартной высоты."""
        safe_append(ov, list(vals) if vals else [None])
        r = ov.max_row
        if not vals:
            ov.row_dimensions[r].height = SPACER_ROW_HEIGHT
            return
        if style == "header":
            style_merged_label(ov, r, 2, header=True)
        elif style == "data":
            if len(vals) == 1:
                style_merged_label(ov, r, 2, header=False)
            else:
                for i in range(1, len(vals) + 1):
                    style_data_cell(ov.cell(r, i))

    _row("Сводный отчёт по проекту «Журнал обхода площадок» — САО г. Москвы")
    ov["A1"].font = Font(bold=True, size=14)
    ov["A1"].alignment = CENTER_WRAP
    _row(f"Снимок на {now_str} (МСК)")
    ov[f"A{ov.max_row}"].alignment = CENTER_WRAP
    _row(
        f"Методика v2 · Europe/Moscow · период "
        f"{stats_filter.date_from:%d.%m.%Y}–{stats_filter.date_to:%d.%m.%Y} · "
        f"сформировано {stats_dashboard.generated_at:%d.%m.%Y %H:%M UTC}"
    )
    ov[f"A{ov.max_row}"].alignment = CENTER_WRAP
    _row()
    _row("Ключевые цифры", style="header")
    _row("Закрыто замечаний за период", total_closures_period, style="data")
    _row("Площадок в системе", total_sites_all, style="data")
    _row("Проверено площадок / не проверено", f"{sites_inspected_all} / {sites_not_inspected_all}", style="data")
    _row("Обходов всего / завершено", f"{total_insp_all} / {total_insp_done_all} (в процессе: {total_insp_in_progress_all})", style="data")
    _row("Обходов без нарушений / с нарушениями", f"{inspections_ok_all} / {inspections_with_defects_all}", style="data")
    _row("Найдено дефектов по чек-листу / замечаний оформлено", f"{total_checklist_defects_all} / {total_iss_all}", style="data")
    _row(
        "Замечаний: в работе / на проверке / на доработке / принято / просрочено",
        f"{total_iss_open_all} / {total_iss_fixed_all} / {total_iss_revision_all} / "
        f"{total_iss_closed_all} / {total_iss_overdue_all}",
        style="data",
    )
    _row("«Проверено» считает площадки, а не записи обходов; «без нарушений» — завершённые обходы без дефектов: «зелёные» обходы входят сюда и никуда не теряются.", style="data")
    _row()
    _row("Требуют внимания в первую очередь", style="header")

    if top_categories:
        cats = ", ".join(f"{cat} ({cnt})" for cat, cnt in top_categories)
        _row(f"• Систематические типы нарушений: {cats}.", style="data")

    if odd_districts:
        _row(f"• Похоже на опечатку/задвоение района при заведении в систему: {', '.join(odd_districts)} — не отдельный реальный район.", style="data")

    _row()
    _row("Лидеры дня по личной активности (обходов начато сегодня)", style="header")
    for _id, full_name, cnt in leaders_today:
        _row(f"   {full_name}", cnt, style="data")

    _row()
    _row("Состав отчёта", style="header")
    for name, desc in [
        ("Задания", "Снимок сейчас: кто отвечает за площадку и когда там были в последний раз"),
        ("Сводка по районам", "Охват (проверено/не проверено), результат обходов (без/с нарушениями) и жизненный цикл замечаний за период, в одной таблице"),
        ("Обходы", "Детальный лог всех обходов: инспектор, площадка, статус, ОК/дефектов/фото"),
        ("Нарушения по чек-листу", "Разбивка нарушений по категориям и конкретным пунктам чек-листа"),
        ("Замечания", "Все зафиксированные замечания с критичностью и статусом"),
        ("Просроченные замечания", "Замечания с истёкшим сроком устранения"),
        ("Динамика", "Активность каждого сотрудника по дням (число отмеченных пунктов чек-листа)"),
    ]:
        _row(name, desc, style="data")

    for col, w in zip("ABCDEF", (55, 30, 20, 15, 15, 15)):
        ov.column_dimensions[col].width = w

    ov["H1"] = "Площадки"
    style_header_cell(ov["H1"], fill=False)
    style_header_cell(ov["I1"], fill=False)
    ov.merge_cells("H1:I1")
    ov["H2"], ov["I2"] = "Проверено", sites_inspected_all
    ov["H3"], ov["I3"] = "Не проверено", sites_not_inspected_all
    for cell in (ov["H2"], ov["I2"], ov["H3"], ov["I3"]):
        style_data_cell(cell)
    ov["H5"] = "Обходы"
    style_header_cell(ov["H5"], fill=False)
    style_header_cell(ov["I5"], fill=False)
    ov.merge_cells("H5:I5")
    ov["H6"], ov["I6"] = "Без нарушений", inspections_ok_all
    ov["H7"], ov["I7"] = "С нарушениями", inspections_with_defects_all
    for cell in (ov["H6"], ov["I6"], ov["H7"], ov["I7"]):
        style_data_cell(cell)
    ov["H9"] = "Замечания"
    style_header_cell(ov["H9"], fill=False)
    style_header_cell(ov["I9"], fill=False)
    ov.merge_cells("H9:I9")
    ov["H10"], ov["I10"] = "В работе", total_iss_open_all
    ov["H11"], ov["I11"] = "Принято", total_iss_closed_all
    for cell in (ov["H10"], ov["I10"], ov["H11"], ov["I11"]):
        style_data_cell(cell)
    if sites_inspected_all + sites_not_inspected_all > 0:
        _pie(ov, "Площадки: проверено / не проверено",
             Reference(ov, min_col=8, min_row=2, max_row=3), Reference(ov, min_col=9, min_row=2, max_row=3), "K2",
             colors=(CHART_GOOD, CHART_MUTED))
    if inspections_ok_all + inspections_with_defects_all > 0:
        _pie(ov, "Обходы: без / с нарушениями",
             Reference(ov, min_col=8, min_row=6, max_row=7), Reference(ov, min_col=9, min_row=6, max_row=7), "K16",
             colors=(CHART_GOOD, CHART_CRITICAL))
    if total_iss_open_all + total_iss_closed_all > 0:
        _pie(ov, "Замечания: в работе / принято",
             Reference(ov, min_col=8, min_row=10, max_row=11), Reference(ov, min_col=9, min_row=10, max_row=11), "K30",
             colors=(CHART_WARNING, CHART_GOOD))

    # Топ категорий нарушений — раньше только одной строкой текста в
    # "Требуют внимания"; та строка остаётся (для быстрого чтения), тут —
    # та же информация в виде графика, чтобы сразу увидеть, где систематика,
    # а где единичный случай.
    if top_categories:
        ov["H13"] = "Категории нарушений"
        style_header_cell(ov["H13"], fill=False)
        style_header_cell(ov["I13"], fill=False)
        ov.merge_cells("H13:I13")
        row = 14
        for cat, cnt in top_categories:
            ov.cell(row, 8, cat)
            ov.cell(row, 9, cnt)
            style_data_cell(ov.cell(row, 8))
            style_data_cell(ov.cell(row, 9))
            row += 1
        _bar(ov, "Топ категорий нарушений",
             Reference(ov, min_col=8, min_row=14, max_row=row - 1),
             Reference(ov, min_col=9, min_row=13, max_row=row - 1), "K44",
             y_title="Нарушений", color=CHART_CRITICAL)
        next_h_row, next_k_row = row + 2, 58
    else:
        next_h_row, next_k_row = 13, 44

    # Топ районов по устранению замечаний — зеркало "Топ категорий
    # нарушений" выше, но по обратной, "хорошей" стороне процесса: не что
    # сломано, а где активнее всего чинят. Явный акцент отчёта на
    # исправления, а не только на учёт находок.
    if district_closures:
        ov.cell(next_h_row, 8, "Районы: закрыто замечаний")
        style_header_cell(ov.cell(next_h_row, 8), fill=False)
        style_header_cell(ov.cell(next_h_row, 9), fill=False)
        ov.merge_cells(start_row=next_h_row, start_column=8, end_row=next_h_row, end_column=9)
        r = next_h_row + 1
        for _did, dname, cnt in district_closures:
            ov.cell(r, 8, dname)
            ov.cell(r, 9, cnt)
            style_data_cell(ov.cell(r, 8))
            style_data_cell(ov.cell(r, 9))
            r += 1
        _bar(ov, "Топ районов по устранению замечаний",
             Reference(ov, min_col=8, min_row=next_h_row + 1, max_row=r - 1),
             Reference(ov, min_col=9, min_row=next_h_row, max_row=r - 1), f"K{next_k_row}",
             y_title="Закрыто", color=CHART_GOOD)

    # Снимок "кто чем занят прямо сейчас", не история за период: это то,
    # что проверяющему нужно открыть первым делом.
    _sheet(wb, "Задания",
        ["Район", "Двор", "Тип площадки", "Назначенный инспектор", "Телефон", "Последний обход", "Статус последнего обхода"],
        assignment_data, [24, 40, 20, 26, 16, 18, 22])
    summary_ws = _sheet(wb, "Сводка по районам",
        ["Район", "Площадок", "Проверено", "Охват %", "Обходов",
         "Без нарушений", "С наруш.", "Выявлено", "Исправлено за период",
         "Устранено за период", "Доработка", "Не устранено", "Просрочено",
         "% устранения из выявленных"],
        summary_rows, [24, 12, 12, 11, 10, 15, 11, 11, 19, 19, 11, 13, 11, 24])
    for row_idx, row in enumerate(summary_rows, start=2):
        for column_idx in (4, 14):
            value = row[column_idx - 1]
            color = "63BE7B" if value >= 100 else "FFD966" if value >= 70 else "F4B183" if value >= 50 else "E06666"
            summary_ws.cell(row_idx, column_idx).fill = PatternFill("solid", fgColor=color)
    summary_ws["O1"], summary_ws["P1"] = "Проверено", "Не проверено"
    for row_idx, row in enumerate(summary_rows, start=2):
        summary_ws.cell(row_idx, 15, row[2])
        summary_ws.cell(row_idx, 16, max(row[1] - row[2], 0))
    total_row_idx = len(summary_rows) + 1
    for column_idx in range(1, 15):
        cell = summary_ws.cell(total_row_idx, column_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="595959")
    summary_ws.column_dimensions["O"].hidden = True
    summary_ws.column_dimensions["P"].hidden = True

    # Сравнение районов — раньше на этом листе не было ни одного графика,
    # хотя это самая насыщенная сравнительная таблица отчёта. Две сгруппи-
    # рованные диаграммы: охват (проверено/не проверено) и результат
    # (без/с нарушениями), по каждому району — та же пара статусных цветов,
    # что и в одноимённых круговых диаграммах на "Обзоре", чтобы смысл
    # цвета не менялся от листа к листу.
    if summary_data:
        n = len(summary_data)
        last_row = 1 + n

        def _grouped_bar(title, col_a, col_b, anchor, colors):
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.title = title
            chart.height, chart.width = 9, 22
            chart.x_axis.axPos = "b"  # см. комментарий в _bar — тот же дефолт-баг openpyxl
            data = Reference(summary_ws, min_col=col_a, max_col=col_b, min_row=1, max_row=last_row)
            cats = Reference(summary_ws, min_col=1, min_row=2, max_row=last_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            for s, color in zip(chart.series, colors):
                s.graphicalProperties.solidFill = color
            summary_ws.add_chart(chart, anchor)

        chart_row = last_row + 3
        _grouped_bar("Охват по районам: проверено / не проверено", 15, 16,
                     f"A{chart_row}", (CHART_GOOD, CHART_MUTED))
        _grouped_bar("Результат по районам: без / с нарушениями", 6, 7,
                     f"A{chart_row + 19}", (CHART_GOOD, CHART_CRITICAL))
    _sheet(wb, "Обходы",
        ["Дата", "Район", "Двор", "Тип площадки", "Инспектор", "Телефон", "Статус", "Начат", "Завершён", "Пунктов ОК", "Дефектов", "Фото", "Комментарий"],
        insp_data, [16, 20, 40, 20, 24, 16, 14, 16, 16, 12, 10, 7, 40])
    _sheet(wb, "Нарушения по чек-листу",
        ["Дата обхода", "Район", "Двор", "Тип площадки", "Категория", "Пункт чек-листа", "Комментарий инспектора", "Инспектор"],
        defect_data, [16, 20, 40, 20, 18, 50, 40, 24])
    _sheet(wb, "Замечания",
        ["Дата", "Район", "Двор", "Заголовок", "Описание", "Критичность", "Статус", "Автор", "Назначено", "Срок", "Устранено"],
        iss_data, [16, 20, 40, 30, 40, 12, 12, 24, 24, 16, 16])
    _sheet(wb, "Просроченные замечания",
        ["Заголовок", "Район", "Двор", "Автор", "Создано", "Срок", "Описание"],
        overdue_rows, [30, 20, 40, 24, 16, 16, 40])

    if dynamics_rows:
        dyn_ws = _sheet(wb, "Динамика",
            ["Дата"] + sorted_inspectors,
            dynamics_rows, [12] + [12] * len(sorted_inspectors))

        # Тренд-график "обходов в день по всему округу" — раньше этот лист
        # был голой матрицей чисел без единой диаграммы. Основная таблица
        # намеренно новые даты сверху (быстро найти сегодня); для линии
        # тренда нужен обратный, хронологический порядок (иначе график
        # читается справа налево) — считаем отдельную компактную таблицу
        # правее основной, а не пересортировываем то, что уже видит
        # пользователь.
        totals_by_day = {dt: sum(day_data[dt].values()) for dt in day_data}
        chrono_dates = sorted(totals_by_day.keys())
        trend_col = len(sorted_inspectors) + 3  # с отступом в 1 колонку от основной таблицы
        dyn_ws.cell(1, trend_col, "Дата")
        dyn_ws.cell(1, trend_col + 1, "Всего обходов")
        style_header_cell(dyn_ws.cell(1, trend_col))
        style_header_cell(dyn_ws.cell(1, trend_col + 1))
        for i, dt in enumerate(chrono_dates, start=2):
            dyn_ws.cell(i, trend_col, dt)
            dyn_ws.cell(i, trend_col + 1, totals_by_day[dt])
            style_data_cell(dyn_ws.cell(i, trend_col))
            style_data_cell(dyn_ws.cell(i, trend_col + 1))
        from openpyxl.utils import get_column_letter
        dyn_ws.column_dimensions[get_column_letter(trend_col)].width = 14
        dyn_ws.column_dimensions[get_column_letter(trend_col + 1)].width = 16
        last_trend_row = 1 + len(chrono_dates)
        if len(chrono_dates) >= 2:
            _line(dyn_ws, "Обходов в день — весь округ",
                  Reference(dyn_ws, min_col=trend_col, min_row=2, max_row=last_trend_row),
                  Reference(dyn_ws, min_col=trend_col + 1, min_row=1, max_row=last_trend_row),
                  dyn_ws.cell(last_trend_row + 3, trend_col).coordinate,
                  y_title="Обходов")

    buf = io.BytesIO()
    wb.save(buf)

    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="journal_export_{stamp}.xlsx"'},
    )
