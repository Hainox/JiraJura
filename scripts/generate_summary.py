"""Сводки:
1) Регистрация по районам — сколько приглашено vs зарегистрировано (по обходам)
2) Общая сводка обходов — KPI, нарушения, замечания
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

USERS_FILE = r"C:\Users\dmitr\Downloads\Учетные записи (1) (1) — с логинами.xlsx"
INSP_FILE = r"C:\Users\dmitr\Downloads\журнал_обходов_2026-08-05.xlsx"
OUT_FILE = r"C:\Users\dmitr\Downloads\Сводки_2026-08-05.xlsx"

# ── Загрузка данных ─────────────────────────────────────────────

wb_users = openpyxl.load_workbook(USERS_FILE)
ws_users = wb_users.active
# Колонки: 1=№, 2=Район, 3=ФИО, 4=Должность, 5=Телефон, 6=Ссылки, 7=РОЛИ, 8=Логин
users_by_district = defaultdict(list)  # район → [(фио, логин, роль), ...]
for r in range(2, ws_users.max_row + 1):
    district = (ws_users.cell(row=r, column=2).value or '').strip()
    fio = (ws_users.cell(row=r, column=3).value or '').strip()
    login = (ws_users.cell(row=r, column=8).value or '').strip()
    role = (ws_users.cell(row=r, column=7).value or '').strip()
    if fio:
        users_by_district[district].append((fio, login, role))

wb_insp = openpyxl.load_workbook(INSP_FILE)

# Лист «Обходы»: 5=Инспектор (ФИО), 7=Статус, 10=Пунктов ОК, 11=Дефектов
ws_insp = wb_insp['Обходы']
inspected_names: set[str] = set()
status_counts = defaultdict(int)
total_ok = 0
total_defects = 0
for r in range(2, ws_insp.max_row + 1):
    fio = (ws_insp.cell(row=r, column=5).value or '').strip()
    status = (ws_insp.cell(row=r, column=7).value or '')
    if fio:
        inspected_names.add(fio)
    if status:
        status_counts[status] += 1
    ok = ws_insp.cell(row=r, column=10).value or 0
    defect = ws_insp.cell(row=r, column=11).value or 0
    total_ok += int(ok)
    total_defects += int(defect)

# Лист «Замечания»: 7=Статус
ws_issues = wb_insp['Замечания']
open_issues = 0
for r in range(2, ws_issues.max_row + 1):
    status = ws_issues.cell(row=r, column=7).value or ''
    if status in ('Открыто', 'Назначено', 'В работе'):
        open_issues += 1

# Лист «Нарушения по чек-листу»
ws_defects = wb_insp['Нарушения по чек-листу']
total_defect_items = ws_defects.max_row - 1  # minus header

# Лист «Сводка по районам» — читаем готовые цифры
ws_summary = wb_insp['Сводка по районам']
district_stats = {}
for r in range(2, ws_summary.max_row + 1):
    name = ws_summary.cell(row=r, column=1).value or ''
    total_sites = ws_summary.cell(row=r, column=2).value or 0
    insp_count = ws_summary.cell(row=r, column=3).value or 0
    issues_created = ws_summary.cell(row=r, column=4).value or 0
    issues_open = ws_summary.cell(row=r, column=6).value or 0
    issues_overdue = ws_summary.cell(row=r, column=7).value or 0
    district_stats[name] = {
        'total_sites': int(total_sites), 'insp_count': int(insp_count),
        'issues_created': int(issues_created), 'issues_open': int(issues_open),
        'issues_overdue': int(issues_overdue),
    }

# ── Расчёт сводки регистрации ───────────────────────────────────

reg_data = []
total_invited_all = 0
total_registered_all = 0
for district in sorted(users_by_district.keys()):
    users = users_by_district[district]
    total_invited = len(users)
    registered = sum(1 for fio, _, _ in users if fio in inspected_names)
    not_registered = total_invited - registered
    pct = round(registered / total_invited * 100, 1) if total_invited > 0 else 0
    total_invited_all += total_invited
    total_registered_all += registered
    reg_data.append((district, total_invited, registered, not_registered, pct))

# ── Запись в Excel ──────────────────────────────────────────────

wb_out = openpyxl.Workbook()

# ── Стили ──
header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2563EB')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
kpi_font = Font(bold=True, size=20, color='111827')
kpi_label_font = Font(size=10, color='6B7280')
green_fill = PatternFill('solid', fgColor='DCFCE7')
red_fill = PatternFill('solid', fgColor='FEE2E2')
yellow_fill = PatternFill('solid', fgColor='FEF3C7')

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_cell(ws, row, col, align=cell_align):
    cell = ws.cell(row=row, column=col)
    cell.alignment = align
    cell.border = thin_border
    return cell

# ════════════════════════════════════════════════════════════════
# Лист 1: Регистрация по районам
# ════════════════════════════════════════════════════════════════
ws1 = wb_out.active
ws1.title = 'Регистрация по районам'

# Заголовок
ws1.merge_cells('A1:F1')
ws1.cell(row=1, column=1, value='Регистрация пользователей по районам — САО г. Москвы')
ws1['A1'].font = Font(bold=True, size=14, color='111827')

# Итоговая строка
ws1.merge_cells('A3:F3')
total_pct = round(total_registered_all / total_invited_all * 100, 1) if total_invited_all > 0 else 0
ws1.cell(row=3, column=1, value=f'ВСЕГО по округу: приглашено {total_invited_all}, зарегистрировано {total_registered_all} ({total_pct}%), не зарегистрировано {total_invited_all - total_registered_all}')
ws1['A3'].font = Font(bold=True, size=12, color='2563EB')

# Таблица
headers1 = ['Район', 'Приглашено', 'Зарег. (есть обходы)', 'Не зарег.', '% охвата']
for c, h in enumerate(headers1, 1):
    ws1.cell(row=5, column=c, value=h)
style_header(ws1, 5, len(headers1))

for i, (district, inv, reg, nreg, pct) in enumerate(reg_data):
    r = 6 + i
    style_cell(ws1, r, 1, Alignment(horizontal='left', vertical='center')).value = district
    style_cell(ws1, r, 2).value = inv
    style_cell(ws1, r, 3).value = reg
    style_cell(ws1, r, 4).value = nreg
    style_cell(ws1, r, 5).value = f'{pct}%'
    # Подсветка
    if pct >= 80:
        for c in range(1, 6):
            ws1.cell(row=r, column=c).fill = green_fill
    elif pct == 0:
        for c in range(1, 6):
            ws1.cell(row=r, column=c).fill = red_fill
    elif pct < 30:
        for c in range(1, 6):
            ws1.cell(row=r, column=c).fill = yellow_fill

# Ширина
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 10

# ════════════════════════════════════════════════════════════════
# Лист 2: Сводка обходов
# ════════════════════════════════════════════════════════════════
ws2 = wb_out.create_sheet('Сводка обходов')
ws2.merge_cells('A1:G1')
ws2.cell(row=1, column=1, value='Сводка обходов — САО г. Москвы (03.08–05.08.2026)')
ws2['A1'].font = Font(bold=True, size=14, color='111827')

# KPI блок
total_insp = sum(status_counts.values())
completed = status_counts.get('Завершён', 0)
in_progress = status_counts.get('В процессе', 0)
issues_found = status_counts.get('Есть нарушения', 0)
critical = status_counts.get('Критический', 0)

kpi_data = [
    ('Всего обходов', total_insp, '1F2937'),
    ('Завершено', completed, '16A34A'),
    ('В процессе', in_progress, 'CA8A04'),
    ('С нарушениями', issues_found, 'EA580C'),
    ('Критических', critical, 'DC2626'),
    ('Пунктов ОК', total_ok, '16A34A'),
    ('Дефектов (чек-лист)', total_defect_items, 'DC2626'),
    ('Замечаний открыто', open_issues, '7C3AED'),
]

for i, (label, value, color) in enumerate(kpi_data):
    col = i * 2 + 1
    # Карточка KPI
    card = ws2.cell(row=3, column=col, value=label)
    card.font = kpi_label_font
    val = ws2.cell(row=4, column=col, value=value)
    val.font = Font(bold=True, size=20, color=color)
    val.alignment = Alignment(horizontal='center')
    ws2.column_dimensions[get_column_letter(col)].width = 18
    ws2.column_dimensions[get_column_letter(col + 1)].width = 3

# Таблица по районам
ws2.merge_cells('A7:G7')
ws2.cell(row=7, column=1, value='По районам')
ws2['A7'].font = Font(bold=True, size=12, color='111827')

headers2 = ['Район', 'Площадок', 'Обходов', 'Замечаний созд.', 'Открыто', 'Просрочено', '% охвата']
for c, h in enumerate(headers2, 1):
    ws2.cell(row=8, column=c, value=h)
style_header(ws2, 8, len(headers2))

row = 9
for district in sorted(district_stats.keys()):
    ds = district_stats[district]
    total = ds['total_sites']
    insp = ds['insp_count']
    pct = round(insp / total * 100, 1) if total > 0 else 0
    style_cell(ws2, row, 1, Alignment(horizontal='left', vertical='center')).value = district
    style_cell(ws2, row, 2).value = total
    style_cell(ws2, row, 3).value = insp
    style_cell(ws2, row, 4).value = ds['issues_created']
    style_cell(ws2, row, 5).value = ds['issues_open']
    style_cell(ws2, row, 6).value = ds['issues_overdue']
    style_cell(ws2, row, 7).value = f'{pct}%'
    row += 1

# Итого по районам
style_cell(ws2, row, 1, Alignment(horizontal='left', vertical='center')).value = 'ИТОГО'
style_cell(ws2, row, 1).font = Font(bold=True)
style_cell(ws2, row, 2).value = sum(d['total_sites'] for d in district_stats.values())
style_cell(ws2, row, 3).value = sum(d['insp_count'] for d in district_stats.values())
style_cell(ws2, row, 4).value = sum(d['issues_created'] for d in district_stats.values())
style_cell(ws2, row, 5).value = sum(d['issues_open'] for d in district_stats.values())
style_cell(ws2, row, 6).value = sum(d['issues_overdue'] for d in district_stats.values())
total_sites_all = sum(d['total_sites'] for d in district_stats.values())
total_insp_all = sum(d['insp_count'] for d in district_stats.values())
style_cell(ws2, row, 7).value = f'{round(total_insp_all/total_sites_all*100,1) if total_sites_all > 0 else 0}%'
for c in range(1, 8):
    ws2.cell(row=row, column=c).font = Font(bold=True)
    ws2.cell(row=row, column=c).border = thin_border

ws2.column_dimensions['A'].width = 24
for c in range(2, 8):
    ws2.column_dimensions[get_column_letter(c)].width = 16

wb_out.save(OUT_FILE)
print(f'✅ Сохранено: {OUT_FILE}')
print(f'\n=== Регистрация ===')
print(f'Всего приглашено: {total_invited_all}')
print(f'Зарегистрировано (есть обходы): {total_registered_all} ({total_pct}%)')
print(f'Не зарегистрировано: {total_invited_all - total_registered_all}')
print(f'\n=== Обходы ===')
print(f'Всего: {total_insp} | Завершено: {completed} | В процессе: {in_progress}')
print(f'С нарушениями: {issues_found} | Критических: {critical}')
print(f'Пунктов ОК: {total_ok} | Дефектов: {total_defect_items} | Замечаний открыто: {open_issues}')
