"""Регрессия: лист "Динамика" в /reports/export.xlsx группирует обходы по
МОСКОВСКИМ суткам, а не по сырой UTC-дате — иначе обход, сделанный в
00:30 по Москве, попадал бы в "вчера" (как и было до фикса). Заодно
регрессия на явный join по district_id — неявный .join(Site) резолвился
через users.id=sites.assigned_inspector_id вместо inspections.site_id,
и фильтр по району на этом листе всегда возвращал пусто. См.
app/routers/reports.py: export_xlsx, day_stats_q."""
import io
import os
import uuid

import psycopg2
import pytest
from openpyxl import load_workbook


SYNC_DB_URL = os.environ["DATABASE_URL"].replace("postgresql+asyncpg://", "postgresql://")


def _exec(sql, params=None):
    conn = psycopg2.connect(SYNC_DB_URL)
    conn.autocommit = True
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or {})
    finally:
        conn.close()


@pytest.mark.asyncio
async def test_dynamics_sheet_buckets_by_moscow_day(client, admin_headers):
    r = await client.get("/api/v1/districts/", headers=admin_headers)
    district_id = r.json()[0]["id"]

    invite = await client.post("/api/v1/auth/invites", json={
        "login": "MskDynInspector", "full_name": "Московсков Инспектор",
        "role": "inspector", "district_id": district_id,
    }, headers=admin_headers)
    complete = await client.post(
        f"/api/v1/auth/invites/{invite.json()['token']}/complete", json={"password": "MskDyn12345"},
    )
    inspector_headers = {"Authorization": f"Bearer {complete.json()['access_token']}"}
    me = await client.get("/api/v1/auth/me", headers=inspector_headers)
    inspector_id = me.json()["id"]

    court_id, site_id = str(uuid.uuid4()), str(uuid.uuid4())
    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES (%(c)s, %(d)s, %(n)s)",
        {"c": court_id, "d": district_id, "n": "Двор МСК-границы"},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "(%(s)s, %(c)s, 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((40.0 55.8,40.001 55.8,40.001 55.801,40.0 55.801,40.0 55.8))', 4326), true)",
        {"s": site_id, "c": court_id},
    )

    # 2026-08-15 21:30:00 UTC = 2026-08-16 00:30:00 MSK (UTC+3) — должно
    # попасть в колонку "2026-08-16", а не "2026-08-15".
    insp_id = str(uuid.uuid4())
    _exec(
        "INSERT INTO inspections (id, site_id, inspector_id, status, created_at) VALUES "
        "(%(i)s, %(s)s, %(u)s, 'completed', '2026-08-15T21:30:00Z')",
        {"i": insp_id, "s": site_id, "u": inspector_id},
    )

    r = await client.get(
        "/api/v1/reports/export.xlsx", params={"district_id": district_id}, headers=admin_headers,
    )
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    assert "Динамика" in wb.sheetnames, (
        f"sheet missing — district filter join is broken again; got {wb.sheetnames}"
    )
    ws = wb["Динамика"]

    header = [c.value for c in ws[1]]
    assert "Московсков Инспектор" in header
    col = header.index("Московсков Инспектор") + 1

    dates_with_count = {
        row[0]: row[col - 1]
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[col - 1]
    }
    assert dates_with_count.get("2026-08-16") == 1, (
        f"expected the inspection under MSK day 2026-08-16, got rows: {dates_with_count}"
    )
    assert "2026-08-15" not in dates_with_count
