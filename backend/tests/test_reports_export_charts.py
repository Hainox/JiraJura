"""Регрессия: /reports/export.xlsx получил дополнительную визуализацию —
графики на "Сводка по районам" (раньше не было ни одного) и "Динамика"
(тренд обходов по округу), плюс единая цветовая палитра вместо цветов
Excel по умолчанию. См. app/routers/reports.py: export_xlsx (_pie/_bar/
_line, CHART_* константы)."""
import io
import os
import uuid

import psycopg2
import pytest
from openpyxl import load_workbook

SYNC_DB_URL = os.environ["DATABASE_URL"].replace("postgresql+asyncpg://", "postgresql://")


def _exec(sql, params=None):
    conn = psycopg2.connect(SYNC_DB_URL)
    conn.autocommit = True
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or {})
    finally:
        conn.close()


def _query_all(sql, params=None):
    conn = psycopg2.connect(SYNC_DB_URL)
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or {})
            return cur.fetchall()
    finally:
        conn.close()


@pytest.mark.asyncio
async def test_district_summary_sheet_has_comparison_charts(client, admin_headers):
    r = await client.get("/api/v1/districts/", headers=admin_headers)
    district_id = r.json()[0]["id"]

    court_id, site_id = str(uuid.uuid4()), str(uuid.uuid4())
    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES (%(c)s, %(d)s, %(n)s)",
        {"c": court_id, "d": district_id, "n": "Двор для графиков"},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "(%(s)s, %(c)s, 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((41.0 55.8,41.001 55.8,41.001 55.801,41.0 55.801,41.0 55.8))', 4326), true)",
        {"s": site_id, "c": court_id},
    )

    r = await client.get("/api/v1/reports/export.xlsx", headers=admin_headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))

    summary_ws = wb["Сводка по районам"]
    assert len(summary_ws._charts) == 2, (
        "expected 2 comparison bar charts (coverage + outcome) on district summary sheet"
    )
    titles = [c.title.tx.rich.p[0].r[0].t for c in summary_ws._charts]
    assert any("Охват" in t for t in titles)
    assert any("Результат" in t for t in titles)
    # Каждая диаграмма — 2 серии (напр. проверено/не проверено), не больше:
    # смешение метрик разного масштаба в одном графике плохо читается.
    for c in summary_ws._charts:
        assert len(c.series) == 2


@pytest.mark.asyncio
async def test_overview_sheet_pies_use_status_colors_not_excel_defaults(client, admin_headers):
    # Явно создаём законченный обход с дефектом — гарантирует ненулевые
    # inspections_with_defects И issues_open, а не полагается на случайно
    # накопленное состояние БД от других тестов сессии (иначе часть из 3
    # круговых диаграмм условно не рисуется, и тест зависит от порядка
    # выполнения других файлов).
    r = await client.get("/api/v1/districts/", headers=admin_headers)
    district_id = r.json()[0]["id"]
    invite = await client.post("/api/v1/auth/invites", json={
        "login": "ChartPieInspector", "full_name": "Пирожковый Инспектор",
        "role": "inspector", "district_id": district_id,
    }, headers=admin_headers)
    complete = await client.post(
        f"/api/v1/auth/invites/{invite.json()['token']}/complete", json={"password": "ChartPie12345"},
    )
    inspector_headers = {"Authorization": f"Bearer {complete.json()['access_token']}"}

    court_id, site_id = str(uuid.uuid4()), str(uuid.uuid4())
    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES (%(c)s, %(d)s, %(n)s)",
        {"c": court_id, "d": district_id, "n": "Двор для пирогов"},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "(%(s)s, %(c)s, 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((43.0 55.8,43.001 55.8,43.001 55.801,43.0 55.801,43.0 55.8))', 4326), true)",
        {"s": site_id, "c": court_id},
    )
    items = _query_all(
        "SELECT id FROM checklist_items WHERE template_id = 'c0000000-0000-0000-0000-000000000001' "
        "AND requires_photo = FALSE ORDER BY sort_order LIMIT 1"
    )
    item_id = str(items[0][0])
    start = await client.post("/api/v1/inspections/", json={"site_id": site_id, "type": "regular"}, headers=inspector_headers)
    insp_id = start.json()["id"]
    done = await client.patch(
        f"/api/v1/inspections/{insp_id}",
        json={"status": "issues_found", "answers": [{"checklist_item_id": item_id, "result": "defect", "comment": "мусор"}]},
        headers=inspector_headers,
    )
    assert done.status_code == 200, done.text

    r = await client.get("/api/v1/reports/export.xlsx", headers=admin_headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ov = wb["Обзор"]

    pies = [c for c in ov._charts if type(c).__name__ == "PieChart"]
    assert len(pies) == 3
    for pie in pies:
        dps = pie.series[0].data_points
        assert dps and len(dps) == 2, "each overview pie should have 2 explicitly colored slices"
        colors = {dp.graphicalProperties.solidFill.srgbClr for dp in dps}
        # Ни одна доля не должна остаться нераскрашенной (осталась бы None).
        assert None not in colors


@pytest.mark.asyncio
async def test_dynamics_sheet_has_okrug_trend_line_chart(client, admin_headers):
    r = await client.get("/api/v1/districts/", headers=admin_headers)
    district_id = r.json()[0]["id"]

    invite = await client.post("/api/v1/auth/invites", json={
        "login": "ChartTrendInspector", "full_name": "Трендовый Инспектор",
        "role": "inspector", "district_id": district_id,
    }, headers=admin_headers)
    complete = await client.post(
        f"/api/v1/auth/invites/{invite.json()['token']}/complete", json={"password": "ChartTrend123"},
    )
    inspector_headers = {"Authorization": f"Bearer {complete.json()['access_token']}"}

    court_id, site_id = str(uuid.uuid4()), str(uuid.uuid4())
    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES (%(c)s, %(d)s, %(n)s)",
        {"c": court_id, "d": district_id, "n": "Двор для тренда"},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "(%(s)s, %(c)s, 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((42.0 55.8,42.001 55.8,42.001 55.801,42.0 55.801,42.0 55.8))', 4326), true)",
        {"s": site_id, "c": court_id},
    )
    start = await client.post("/api/v1/inspections/", json={"site_id": site_id, "type": "regular"}, headers=inspector_headers)
    insp_id = start.json()["id"]
    _exec("UPDATE inspections SET status='completed', created_at = now() - interval '1 day' WHERE id = %(i)s", {"i": insp_id})

    court_id2, site_id2 = str(uuid.uuid4()), str(uuid.uuid4())
    _exec(
        "INSERT INTO courtyards (id, district_id, name) VALUES (%(c)s, %(d)s, %(n)s)",
        {"c": court_id2, "d": district_id, "n": "Двор для тренда 2"},
    )
    _exec(
        "INSERT INTO sites (id, courtyard_id, type, area_m2, geometry, is_active) VALUES "
        "(%(s)s, %(c)s, 'Детская площадка', 100, "
        "ST_GeomFromText('POLYGON((42.1 55.8,42.101 55.8,42.101 55.801,42.1 55.801,42.1 55.8))', 4326), true)",
        {"s": site_id2, "c": court_id2},
    )
    start2 = await client.post("/api/v1/inspections/", json={"site_id": site_id2, "type": "regular"}, headers=inspector_headers)
    insp_id2 = start2.json()["id"]
    _exec("UPDATE inspections SET status='completed', created_at = now() WHERE id = %(i)s", {"i": insp_id2})

    r = await client.get("/api/v1/reports/export.xlsx", headers=admin_headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    assert "Динамика" in wb.sheetnames
    dyn = wb["Динамика"]
    line_charts = [c for c in dyn._charts if type(c).__name__ == "LineChart"]
    assert len(line_charts) == 1
    assert len(line_charts[0].series) == 1

    # Мини-таблица тренда — правее основной, "Всего обходов" по возрастанию дат.
    header = [c.value for c in dyn[1]]
    assert "Всего обходов" in header
    col = header.index("Всего обходов") + 1
    date_col = col - 1
    dates = [row[0] for row in dyn.iter_rows(min_row=2, min_col=date_col, max_col=date_col, values_only=True) if row[0]]
    assert dates == sorted(dates), "trend table must be chronological (ascending), not newest-first"
