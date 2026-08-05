"""Сравнение Excel (приглашённые) с реальной БД (83 пользователя).
Вход: Excel + текстовый список из админки (скопирован из сообщения).
Выход: обновлённая сводка + список расхождений.
"""
import openpyxl
from collections import defaultdict

USERS_FILE = r"C:\Users\dmitr\Downloads\Учетные записи (1) (1) — с логинами.xlsx"
OUT_FILE = r"C:\Users\dmitr\Downloads\Сводки_2026-08-05.xlsx"

# ── 83 пользователя из БД (из админки): ФИО, Логин, Роль, Район, Телефон, Статус ──
db_raw = """1234567890	1234567890	Инспектор	Аэропорт	—	активен
12345678901	12345678901	Проверяющий	Бескудниковский	—	активен
Администратор	admin	Админ	—	—	активен
Андреева Надежда Валерьевна	AndreevaNV	Проверяющий	Бескудниковский	—	активен
Асеков Азамат Алимсултанович	AsekovAA	Проверяющий	Дмитровский	—	активен
Балицкий Павел Александрович	BalitskyPA	Инспектор	Левобережный	—	активен
Барков Олег Александрович	BarkovOA	Инспектор	Левобережный	—	активен
Бойматов Бехзод Курбанович	BoymatovBK1	Инспектор	Головинский	—	активен
Бородинова Мария Константиновна	BorodinovaMK	Проверяющий	Беговой	—	активен
Братская Александра Сергеевна	BratskayaAS	Инспектор	Восточное Дегунино	—	активен
Бузенкова Елена Сергеевна	BuzenkovaES	Инспектор	Тимирязевский	—	активен
Бурмистров Владислав Геннадьевич	BurmistrovVG	Инспектор	Беговой	—	активен
Варжин Артем Игоревич	VarzhinAI	Инспектор	Коптево	—	активен
Василевский Никита Сергеевич	VasilevskyNS	Инспектор	Дмитровский	—	активен
Гевенов Левон Акопович	GevenovLA	Инспектор	Головинский	—	активен
Гнедарева Марина Николаевна	GnedarevaMN1	Проверяющий	Войковский	—	активен
Громов Владимир Владимирович	GromovVV	Проверяющий	Восточное Дегунино	—	активен
Гуринова Елена Геннадьевна	GurinovaEG	Инспектор	Беговой	—	активен
Джабраилова Елена Михайловна	DzhabrailovaEM	Инспектор	Тимирязевский	—	активен
Джалавханов Ибрагим Пируллахович	DzhalavhanovIP	Инспектор	Хорошевский	—	активен
Евдокимова Галина Николаевна	EvdokimovaGN	Инспектор	Войковский	—	активен
Забабурин Павел Иванович	ZababurinPI	Инспектор	Восточное Дегунино	—	активен
Залогина Марина Мансуровна	ZaloginaMM	Инспектор	Молжаниновский	—	активен
Зейналов Сабир Алыханович	ZeynalovSA	Инспектор	Хорошевский	—	активен
Исайчев Сергей Михайлович	IsaychevSM	Инспектор	Хорошевский	—	активен
Калачев Алексей Александрович	KalachevAA	Инспектор	Беговой	—	активен
Калашникова Ирина Валентиновна	KalashnikovaIV	Проверяющий	Савеловский	—	активен
Карпачев Александр Владимирович	KarpachevAV	Проверяющий	Дмитровский	—	активен
Керимов Радж Шахин Оглы	deleted_9c4cb5dc	Проверяющий	Ховрино	—	отключён
Керимов Радж Шахин Оглы	KerimovRS2	Проверяющий	Ховрино	—	активен
Кипер Людмила Александровна	KiperLA	Инспектор	Головинский	—	активен
Кирсанов Андрей Константинович	KirzhanovAK	Инспектор	Молжаниновский	—	активен
Клевцов Денис Сергеевич	KlevtsovDS1	Инспектор	Головинский	—	активен
Колесьянкина Надежда Викторовна	KolesyankinaNV	Инспектор	Беговой	—	активен
Коротков Владимир Игоревич	KorotkovVI	Инспектор	Аэропорт	—	активен
Корсаков Алексей Васильеви	KorsakovaFV	Инспектор	Молжаниновский	—	активен
Кузнецов Павел Владимирович	KuznetsovPV	Инспектор	Молжаниновский	—	активен
Кузьмин Денис Юрьевич	KuzminDYu	Инспектор	Восточное Дегунино	—	активен
Ломакин Дмитрий Александрович	LomakinDA	Инспектор	Восточное Дегунино	—	активен
Лукаш Ольга Николаевна	LukashON	Инспектор	Головинский	—	активен
Магомедов Мурад Исмаилович	MagomedovMI	Инспектор	Беговой	—	активен
Магомедов Надыр Н.	MagomedovNN	Инспектор	Дмитровский	—	активен
Масюкевич Елена	MasyukevichE	Инспектор	Восточное Дегунино	—	активен
Мичурин Дмитрий Дмитриевич	MichurinDD2	Проверяющий	Беговой	—	активен
Мичурин Дмитрий Дмитриевич	MichurinDD1	Инспектор	Беговой	—	активен
Мичурин Дмитрий Дмитриевич	MichurinDD	Админ	—	+79256590813	активен
Монахов Владислав Юрьевич	MonakhovVYu	Инспектор	Коптево	—	активен
Мосин Сергей Владимирович	MosinSV	Инспектор	Молжаниновский	—	активен
Наумов Артем Сергеевич	NaumovAS	Инспектор	Коптево	—	активен
Нефедов Константин Игоревич	NefedovKI	Инспектор	Коптево	—	активен
Новичков Иван Романович	NovichkovIR	Инспектор	Головинский	—	активен
Нурманов Руслан	NurmanovR	Инспектор	Головинский	—	активен
Осмоловский Алексей Геннадьевич	OsmolovskyAG	Инспектор	Бескудниковский	—	активен
Перепелкина Наталья Владимировна	PerepelkinaNV	Инспектор	Восточное Дегунино	—	активен
Пожидаева Мария Владимировна	PozhidaevaMV	Инспектор	Восточное Дегунино	—	активен
Пугач Юрий Николаевич	PugachYuN	Инспектор	Хорошевский	—	активен
Романенко Михаил Александрович	RomanenkoMA	Инспектор	Головинский	—	активен
Савченко Владимир Александрович	SavchenkoVA	Инспектор	Левобережный	—	активен
Самарина Наталья Вячеславовна	SamarinaNV	Проверяющий	Западное Дегунино	—	активен
Самохвалов Владимир Яковлевич	SamokhvalovVYa	Инспектор	Дмитровский	—	активен
Сафронов Егор Алексеевич	SafronovEA	Инспектор	Аэропорт	—	активен
Семенов Михайл Юрьевич	SemenovMYu	Инспектор	Хорошевский	—	активен
Сень Светлана Анатольевна	SenSA	Инспектор	Дмитровский	—	активен
Скуратов Иван Сергеевич	SkuratovIS	Инспектор	Аэропорт	—	активен
Соколов Сергей Сергеевич	SokolovSS	Инспектор	Восточное Дегунино	—	активен
Спирина Анна Николаевна	SpirinaAN	Инспектор	Тимирязевский	—	активен
Стахова Светлана Игоревна	StakhovaSI	Инспектор	Молжаниновский	—	активен
Степанчук Алла Валентиновна	StepanchukAV	Проверяющий	Молжаниновский	—	активен
Сумакова Наталья Александровна	SumakovaNA	Инспектор	Восточное Дегунино	—	активен
Сухин Вячеслав Станиславович	SukhinVS	Инспектор	Восточное Дегунино	—	активен
Таланова Анна Сергеевна	TalanovaAS	Проверяющий	Сокол	—	активен
Тарасик Вадим Артурович	TarasikVA	Инспектор	Сокол	—	активен
Токарева Елена Евгеньевна	TokarevaEE	Инспектор	Восточное Дегунино	—	активен
Ходаковская Юлия Васильевна	KhodakovskayaYuV	Инспектор	Коптево	—	активен
Хорчев Анатолий Михайлович	HorchevAM	Инспектор	Молжаниновский	—	активен
Чекалкина Светлана Михайловна	ChekalkinaSM	Инспектор	Восточное Дегунино	—	активен
Черкашин Михаил Васильевич	CherkashinMV	Инспектор	Коптево	—	активен
Чиликина Татьяна Николаевна	ChilikinaTN	Инспектор	Восточное Дегунино	—	активен
Чумичкина Екатерина Валерьевна	ChumichkinaEV1	Инспектор	Тимирязевский	—	активен
Шазамова Алсу Рифатовна	ShazamovaAR	Инспектор	Восточное Дегунино	—	активен
Шкварин Дмитрий Сергеевич	ShkvarinDS	Инспектор	Восточное Дегунино	—	активен
Шокиров Шокирджон	ShokirovSh	Инспектор	Головинский	—	активен
Юркова Арина Владиславовна	YurkovaAV1	Проверяющий	Тимирязевский	—	активен"""

# Парсим DB-данные
db_users = {}  # login → (fio, role, district, status)
db_by_fio = {}  # fio → login
for line in db_raw.strip().split('\n'):
    parts = [p.strip() for p in line.split('\t')]
    if len(parts) >= 6:
        fio, login, role, district, phone, status = parts[0], parts[1], parts[2], parts[3], parts[4], parts[5]
        if status == 'активен' and login != 'admin' and login != '1234567890' and login != '12345678901' and not login.startswith('deleted_'):
            db_users[login] = (fio, role, district, status)
            db_by_fio[fio] = login

print(f"DB активных (без admin/мусора): {len(db_users)}")

# Читаем Excel
wb = openpyxl.load_workbook(USERS_FILE)
ws = wb.active
invited = {}  # login → (fio, district, role)
for r in range(2, ws.max_row + 1):
    fio = (ws.cell(row=r, column=3).value or '').strip()
    login = (ws.cell(row=r, column=8).value or '').strip()
    district = (ws.cell(row=r, column=2).value or '').strip()
    role = (ws.cell(row=r, column=7).value or '').strip()
    if login and fio:
        invited[login] = (fio, district, role)

print(f"Excel приглашённых: {len(invited)}")

# Сравнение
excel_logins = set(invited.keys())
db_logins = set(db_users.keys())

matched = excel_logins & db_logins
only_in_excel = excel_logins - db_logins  # приглашены но не зареганы
only_in_db = db_logins - excel_logins      # в БД но не в Excel (возможно, расхождения)

print(f"\nСовпало (логин есть и в Excel, и в БД): {len(matched)}")
print(f"Есть в Excel, НЕТ в БД (не зарегистрировались): {len(only_in_excel)}")
print(f"Есть в БД, НЕТ в Excel (не были в списке приглашённых?): {len(only_in_db)}")

# ── Анализ расхождений ──
print("\n=== Есть в БД, НЕТ в Excel ===")
for login in sorted(only_in_db):
    fio, role, district, status = db_users[login]
    print(f"  {login:<22} | {fio:<35} | {role:<14} | {district}")

print("\n=== Расхождения по логинам (совпали по ФИО, но разные логины) ===")
# Ищем совпадения по ФИО
db_fio_set = set(db_by_fio.keys())
excel_fio_set = set(fio for fio, _, _ in invited.values())
fio_matched = db_fio_set & excel_fio_set
for fio in sorted(fio_matched):
    db_login = db_by_fio[fio]
    # Найти excel_login по ФИО
    excel_login = None
    for l, (f, _, _) in invited.items():
        if f == fio:
            excel_login = l
            break
    if db_login != excel_login:
        print(f"  ФИО: {fio}")
        print(f"    DB-логин:     {db_login}")
        print(f"    Excel-логин:  {excel_login}")

# ── Порайонная сводка ──
district_reg = defaultdict(lambda: {'invited': 0, 'registered': 0, 'not_reg': []})

# Проходим по Excel: каждый приглашённый
for login, (fio, district, role) in invited.items():
    district_reg[district]['invited'] += 1
    if login in db_users:
        district_reg[district]['registered'] += 1
    else:
        district_reg[district]['not_reg'].append(fio)

print("\n=== ПОРАЙОННАЯ СВОДКА (реальные данные) ===")
total_inv = 0
total_reg = 0
for district in sorted(district_reg.keys()):
    d = district_reg[district]
    inv = d['invited']
    reg = d['registered']
    nreg = inv - reg
    pct = round(reg / inv * 100, 1) if inv > 0 else 0
    total_inv += inv
    total_reg += reg
    bar = '█' * int(pct / 10) + '░' * (10 - int(pct / 10))
    print(f"  {district:<22} | {inv:>3} пригл. | {reg:>3} зарег. | {nreg:>3} нет | {pct:>5.1f}% {bar}")

total_pct = round(total_reg / total_inv * 100, 1) if total_inv > 0 else 0
print(f"\n  {'ВСЕГО':22} | {total_inv:>3} пригл. | {total_reg:>3} зарег. | {total_inv - total_reg:>3} нет | {total_pct:>5.1f}%")

# ── НЕЗАРЕГИСТРИРОВАННЫЕ по районам ──
print("\n=== НЕЗАРЕГИСТРИРОВАННЫЕ ===")
for district in sorted(district_reg.keys()):
    d = district_reg[district]
    if d['not_reg']:
        print(f"\n{district} ({len(d['not_reg'])} чел.):")
        for fio in d['not_reg']:
            print(f"  • {fio}")
